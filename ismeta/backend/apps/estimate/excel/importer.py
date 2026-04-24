"""Excel import — парсит .xlsx, создаёт/обновляет items через EstimateService.

E7 base + E7.2 smart auto-detect: нечёткий маппинг заголовков, одна колонка
цены, skip строк-итогов.
"""

import logging
import re
import zipfile
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from apps.estimate.models import Estimate, EstimateItem, EstimateSection
from apps.estimate.services.estimate_service import EstimateService
from apps.estimate.services.markup_service import recalc_estimate_totals

logger = logging.getLogger(__name__)


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)


def _to_decimal(val, default=Decimal("0")) -> Decimal:
    if val is None:
        return default
    try:
        return Decimal(str(val)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return default


# ---------------------------------------------------------------------------
# Auto-detect column mapping
# ---------------------------------------------------------------------------

# Нечёткий маппинг: ключ → набор возможных заголовков (normalised: lower,
# strip trailing punctuation, ё→е).
#
# TD-02 (DEV-BACKLOG #12): поля UI-04 tech_specs — Модель/Производитель/
# Бренд/Примечание/Система — должны распознаваться при импорте. Наряду с
# экспортом (round-trip) это закрывает GAP «правка в Excel теряется».
#
# После E15.05 it2 brand и manufacturer — два разных поля:
#   brand        = торговая марка (Корф, IEK, Fujitsu)
#   manufacturer = конкретный завод-изготовитель (ООО «КОРФ», АО «ДКС»)
# Поэтому Производитель → manufacturer, а не brand как было раньше.
# Алиасы записаны в normalised-виде: lower, ё→е, spaces+punctuation удалены.
# Нормализатор header'ов при импорте применяет те же правила, поэтому «Ед.изм.»,
# «Ед. изм», «Ед изм» и «ЕД.ИЗМ» все становятся «едизм».
_HEADER_ALIASES = {
    "name": {"наименование", "название", "позиция", "описание", "наим", "имя"},
    "unit": {"ед", "единица", "едизм", "единицаизмерения", "еи"},
    "quantity": {"колво", "количество", "кво", "кол", "количествошт"},
    "equipment_price": {"ценаоборуд", "ценаоборудования", "оборудование"},
    "material_price": {"ценамат", "ценаматериалов", "материалы",
                       "ценаматзакуп", "ценаматериаловзакуп"},
    "work_price": {"ценаработ", "работы", "монтаж", "стоимостьработ",
                   "ценаработзакуп"},
    "price": {"цена", "стоимость", "ценазаед", "ценаед"},
    "row_id": {"rowid", "idстроки", "идентификатор"},
    "model_name": {"модель", "model", "марка", "артикул", "тип", "sku",
                   "обозначение", "обозначениедокумента"},
    "brand": {"бренд", "brand"},
    "manufacturer": {"производитель", "изготовитель", "поставщик",
                     "вендор", "manufacturer", "vendor"},
    "comments": {"примечание", "примечания", "комментарий", "комментарии",
                 "заметка", "notes", "comment", "comments"},
    "system": {"система", "контур", "system"},
}


def _normalize_header(val) -> str:
    """Normalise header cell: lower, ё→е, удалить пробелы и пунктуацию."""
    if val is None:
        return ""
    s = str(val).strip().lower().replace("ё", "е")
    # Удаляем пробелы, точки, запятые, двоеточия, скобки, дефисы, подчёркивания.
    return re.sub(r"[\s.,:;()\-_]+", "", s)

# Строки-итоги для пропуска
_SKIP_PATTERNS = re.compile(
    r"^\s*(итого.*|всего.*|total.*|subtotal.*|итог|сумма)\s*$",
    re.IGNORECASE,
)


@dataclass
class ColumnMapping:
    """Маппинг колонок Excel → полей ISMeta."""

    name: int | None = None
    unit: int | None = None
    quantity: int | None = None
    equipment_price: int | None = None
    material_price: int | None = None
    work_price: int | None = None
    price: int | None = None  # единая колонка цены
    row_id: int | None = None
    # TD-02: tech_specs поля UI-04. Складываются в item.tech_specs dict.
    model_name: int | None = None
    brand: int | None = None
    manufacturer: int | None = None
    comments: int | None = None
    system: int | None = None


# Набор ключей маппинга, уходящих в tech_specs (а не в top-level item field).
_TECH_SPECS_KEYS: tuple[str, ...] = (
    "model_name", "brand", "manufacturer", "comments", "system",
)


def _detect_columns(header_row: list) -> ColumnMapping:
    """Auto-detect колонок по заголовкам первой строки."""
    mapping = ColumnMapping()
    for idx, cell_val in enumerate(header_row):
        if cell_val is None:
            continue
        val = _normalize_header(cell_val)
        if not val:
            continue
        for field_name, aliases in _HEADER_ALIASES.items():
            if val in aliases:
                # Не переопределяем — первый матч побеждает (чтобы не сломать
                # существующий порядок алиасов при перекрывающихся ключах).
                if getattr(mapping, field_name) is None:
                    setattr(mapping, field_name, idx)
                break
    return mapping


def _fallback_mapping() -> ColumnMapping:
    """Fallback: стандартный порядок колонок ISMeta."""
    return ColumnMapping(
        name=0, unit=1, quantity=2,
        equipment_price=3, material_price=4, work_price=5, row_id=6,
    )


def _get_val(values: list, idx: int | None, default=None):
    if idx is None or idx >= len(values):
        return default
    return values[idx]


def _is_skip_row(name_val) -> bool:
    """Проверить строку-итог для пропуска."""
    if not name_val:
        return False
    return bool(_SKIP_PATTERNS.match(str(name_val).strip()))


# ---------------------------------------------------------------------------
# Main importer
# ---------------------------------------------------------------------------


def import_estimate_xlsx(estimate_id, workspace_id, file) -> ImportResult:
    """Парсит .xlsx, создаёт/обновляет позиции.

    Smart auto-detect: определяет колонки по заголовкам, поддерживает одну
    колонку «Цена», пропускает строки-итоги.
    """
    result = ImportResult()

    try:
        wb = load_workbook(file, read_only=True)
    except (InvalidFileException, KeyError, zipfile.BadZipFile) as e:
        result.errors.append(f"Невалидный файл: {e}")
        return result

    ws = wb.active
    if ws is None:
        result.errors.append("Нет активного листа")
        return result

    estimate = Estimate.objects.get(id=estimate_id, workspace_id=workspace_id)

    # Читаем все строки (read_only → iter)
    all_rows = list(ws.iter_rows(values_only=False))
    if not all_rows:
        return result

    # Auto-detect колонок по первой строке
    header_values = [cell.value for cell in all_rows[0]]
    mapping = _detect_columns(header_values)
    has_headers = mapping.name is not None

    if not has_headers:
        mapping = _fallback_mapping()
    # Первая строка всегда пропускается (заголовки или нераспознанные заголовки)
    data_rows = all_rows[1:]

    current_section = None
    sort_order = 0

    with transaction.atomic():
        create_batches: dict[str, list[dict]] = {}

        for row_idx, row in enumerate(data_rows, start=2 if has_headers else 1):
            values = [cell.value for cell in row]

            if not any(values):
                continue

            name_val = _get_val(values, mapping.name)

            # Skip итогов
            if _is_skip_row(name_val):
                result.skipped += 1
                continue

            # Section detection: жирный текст в первой ячейке, остальные пустые
            cell_0 = row[mapping.name or 0]
            non_name_values = [v for i, v in enumerate(values) if i != (mapping.name or 0)]
            is_section = (
                cell_0.value
                and cell_0.font
                and cell_0.font.bold
                and not any(non_name_values[:5])
            )

            if is_section:
                section_name = str(name_val).strip()
                if _is_skip_row(section_name):
                    result.skipped += 1
                    continue
                current_section, _ = EstimateSection.objects.get_or_create(
                    estimate=estimate,
                    workspace_id=workspace_id,
                    name=section_name,
                    defaults={"sort_order": sort_order},
                )
                sort_order += 1
                continue

            if current_section is None:
                current_section, _ = EstimateSection.objects.get_or_create(
                    estimate=estimate,
                    workspace_id=workspace_id,
                    name="Импорт",
                    defaults={"sort_order": 0},
                )

            if not name_val or not str(name_val).strip():
                result.errors.append(f"Строка {row_idx}: пустое наименование")
                continue

            unit_val = _get_val(values, mapping.unit)
            unit = str(unit_val).strip() if unit_val else "шт"
            quantity = _to_decimal(_get_val(values, mapping.quantity))

            # Цены: если есть отдельные колонки — используем, иначе price → material_price
            equip_price = _to_decimal(_get_val(values, mapping.equipment_price))
            mat_price = _to_decimal(_get_val(values, mapping.material_price))
            work_price = _to_decimal(_get_val(values, mapping.work_price))

            if mapping.price is not None and mat_price == 0 and equip_price == 0:
                mat_price = _to_decimal(_get_val(values, mapping.price))

            row_id_val = _get_val(values, mapping.row_id)
            row_id = str(row_id_val).strip() if row_id_val else None

            if quantity < 0:
                result.errors.append(f"Строка {row_idx}: отрицательное количество ({quantity})")
                continue

            # TD-02: читаем колонки tech_specs (UI-04) если они есть.
            spec_overrides: dict[str, str] = {}
            for key in _TECH_SPECS_KEYS:
                col_idx = getattr(mapping, key)
                if col_idx is None:
                    continue
                raw = _get_val(values, col_idx)
                if raw is None:
                    continue
                s = str(raw).strip()
                if s:
                    spec_overrides[key] = s

            sort_order += 1
            data = {
                "name": str(name_val).strip(),
                "unit": unit,
                "quantity": quantity,
                "equipment_price": equip_price,
                "material_price": mat_price,
                "work_price": work_price,
                "sort_order": sort_order,
            }

            if row_id:
                existing = EstimateItem.all_objects.filter(
                    row_id=row_id, estimate_id=estimate_id, workspace_id=workspace_id
                ).first()
                if existing:
                    # Merge с существующим tech_specs, чтобы не потерять ключи
                    # которых нет в Excel-шаблоне (dimensions, power_kw и т.д.).
                    if spec_overrides:
                        merged = dict(existing.tech_specs or {})
                        merged.update(spec_overrides)
                        data["tech_specs"] = merged
                    EstimateService.update_item(
                        existing.id, workspace_id, existing.version, data
                    )
                    result.updated += 1
                    continue

            if spec_overrides:
                data["tech_specs"] = spec_overrides

            create_batches.setdefault(current_section.id, []).append(data)

        for sec_id, batch in create_batches.items():
            sec = EstimateSection.objects.get(id=sec_id)
            result.created += EstimateService.bulk_create_items(sec, estimate, workspace_id, batch)

        recalc_estimate_totals(estimate_id, workspace_id)

    return result
