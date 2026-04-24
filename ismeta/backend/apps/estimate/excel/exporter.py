"""Excel export — формат внутренней сметы Августа.

Sheet 1 "Смета": разделы-разделители + строки + итоги + скрытые row_id/hash.
Sheet 2 "Агрегаты": summary.

TD-02 (#28): колонка «Модель» пропадала в экспорте (только name). После
UI-04 tech_specs хранит model_name / brand / manufacturer / comments / system
— все они должны попадать в Excel для двустороннего round-trip
(DEV-BACKLOG #12).
"""

import hashlib
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.estimate.models import Estimate, EstimateItem, EstimateSection

HEADER_FONT = Font(bold=True, size=11)
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)

# TD-02: расширяем набор колонок под поля UI-04 из tech_specs.
# Порядок: базовые идентификаторы → ТТХ → цены → финансы → примечание →
# скрытые поля round-trip.
COLUMNS = [
    ("№", 6),
    ("Наименование", 40),
    ("Модель", 18),
    ("Производитель", 18),
    ("Бренд", 14),
    ("Система", 16),
    ("Ед.изм.", 10),
    ("Кол-во", 10),
    ("Цена оборуд.", 14),
    ("Цена мат. (закуп)", 16),
    ("Цена мат. (продажа)", 18),
    ("Цена работ (закуп)", 16),
    ("Цена работ (продажа)", 18),
    ("Итого", 14),
    ("Примечание", 40),
    ("row_id", 0),  # скрытый
    ("row_hash", 0),  # скрытый
]


def _row_hash(item) -> str:
    """SHA256-хэш значимых полей для Excel round-trip (ADR-0013)."""
    data = f"{item.name}|{item.unit}|{item.quantity}|{item.equipment_price}|{item.material_price}|{item.work_price}|{item.total}"
    return hashlib.sha256(data.encode()).hexdigest()[:16]


def _spec_str(item, key: str) -> str:
    """Безопасное чтение строкового поля tech_specs (JSONField)."""
    specs = item.tech_specs or {}
    val = specs.get(key)
    return str(val) if val is not None else ""


def export_estimate_xlsx(estimate_id, workspace_id) -> io.BytesIO:
    """Генерирует .xlsx и возвращает BytesIO."""
    estimate = Estimate.objects.get(id=estimate_id)
    sections = EstimateSection.objects.filter(estimate=estimate).order_by("sort_order")

    wb = Workbook()

    # --- Sheet 1: Смета ---
    ws = wb.active
    ws.title = "Смета"

    # Header row
    for col_idx, (title, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        if width > 0:
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        else:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    row_num = 2
    item_counter = 0

    for section in sections:
        # Section header row
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(COLUMNS))
        cell = ws.cell(row=row_num, column=1, value=section.name)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(horizontal="left")
        row_num += 1

        items = EstimateItem.objects.filter(
            section=section, estimate_id=estimate_id, workspace_id=workspace_id
        ).order_by("sort_order")

        for item in items:
            item_counter += 1
            values = [
                item_counter,
                item.name,
                _spec_str(item, "model_name"),
                _spec_str(item, "manufacturer"),
                _spec_str(item, "brand"),
                _spec_str(item, "system"),
                item.unit,
                float(item.quantity),
                float(item.equipment_price),
                float(item.material_price),
                float(item.material_total / item.quantity) if item.quantity else 0,
                float(item.work_price),
                float(item.work_total / item.quantity) if item.quantity else 0,
                float(item.total),
                _spec_str(item, "comments"),
                str(item.row_id),
                _row_hash(item),
            ]
            for col_idx, val in enumerate(values, 1):
                ws.cell(row=row_num, column=col_idx, value=val)
            row_num += 1

    # Totals row
    row_num += 1
    ws.cell(row=row_num, column=1, value="ИТОГО").font = TOTAL_FONT
    # TD-02: колонка «Итого» переехала с 10 → 14 после вставки Модель/
    # Производитель/Бренд/Система.
    ws.cell(row=row_num, column=14, value=float(estimate.total_amount)).font = TOTAL_FONT

    # --- Sheet 2: Агрегаты ---
    ws2 = wb.create_sheet("Агрегаты")
    agg_data = [
        ("Итого оборудование", float(estimate.total_equipment)),
        ("Итого материалы", float(estimate.total_materials)),
        ("Итого работы", float(estimate.total_works)),
        ("Итого", float(estimate.total_amount)),
        ("Человеко-часы", float(estimate.man_hours)),
        ("Прибыльность %", float(estimate.profitability_percent)),
        ("Аванс", float(estimate.advance_amount)),
        ("Сроки (дни)", estimate.estimated_days),
    ]
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20
    for i, (label, val) in enumerate(agg_data, 1):
        ws2.cell(row=i, column=1, value=label).font = HEADER_FONT
        ws2.cell(row=i, column=2, value=val)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
