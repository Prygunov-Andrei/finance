"""Тесты demo-сметы формата ОВ2."""

import io
import uuid

import pytest
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.styles import Font

from apps.estimate.excel.importer import import_estimate_xlsx
from apps.estimate.models import Estimate, EstimateItem, EstimateSection
from apps.workspace.models import Workspace

User = get_user_model()


def _make_demo_xlsx():
    """Создаёт demo xlsx с ≥25 позициями по структуре ОВ2."""
    wb = Workbook()
    ws = wb.active
    BOLD = Font(bold=True)
    ws.append(["Наименование", "Ед.изм.", "Кол-во", "Цена оборуд.", "Цена мат.", "Цена работ"])

    def sec(name):
        ws.append([name])
        for c in ws[ws.max_row]:
            c.font = BOLD

    def item(name, unit, qty, eq=0, mat=0, work=0):
        ws.append([name, unit, qty, eq, mat, work])

    sec("Вытяжная вентиляция — Жилая часть")
    item("Дефлектор Цаги ф355-оц-фл", "шт", 58, 3200, 0, 1500)
    item("Узел прохода УП1 ф355-1000", "шт", 58, 4800, 0, 2500)
    item("Воздуховод прямоугольный 150x100", "м.п.", 3135, 0, 520, 350)
    item("Воздуховод прямоугольный 200x200", "м.п.", 850, 0, 680, 400)
    item("Воздуховод прямоугольный 300x300", "м.п.", 759, 0, 920, 500)
    item("Воздуховод прямоугольный 350x350", "м.п.", 808, 0, 1050, 550)
    item("Изоляция противопожарная EI30 PRO-MBOP-VENT", "м.кв.", 4900, 0, 280, 120)

    sec("Вытяжная вентиляция — МОП")
    item("Зонт выбросной 400x200", "шт", 1, 8500, 0, 3500)
    item("Вентилятор канальный WNK 100/1 Корф", "шт", 10, 28500, 0, 8500)
    item("Вентилятор канальный WNK 200/1 Корф", "шт", 2, 42000, 0, 12000)
    item("Огнезадерживающий клапан EI60 OKL-2K-60-ф160", "шт", 2, 12500, 0, 3500)
    item("Решётка АМН-300x100", "шт", 4, 1800, 0, 650)
    item("Обратный клапан ф100", "шт", 10, 1200, 0, 450)
    item("Решётка вытяжная круглая БСК ф100", "шт", 11, 650, 0, 350)
    item("Воздуховод круглый ф100 (МОП)", "м.п.", 1245, 0, 450, 300)
    item("Воздуховод круглый ф250", "м.п.", 248, 0, 750, 420)

    sec("Противодымная вентиляция")
    item("Вентилятор дымоудаления KLR-DU-400-80H", "шт", 3, 185000, 0, 25000)
    item("Вентилятор дымоудаления KLR-DU-400-63H", "шт", 2, 168000, 0, 22000)
    item("Вентилятор подпора осевой KSO 50-3x30", "шт", 3, 45000, 0, 15000)
    item("Моноблочная установка приточная UTR 50-25", "шт", 3, 280000, 0, 35000)
    item("Вентилятор подпора крышный KSP 80-11x30", "шт", 2, 95000, 0, 22000)

    sec("Клапаны противодымные")
    item("Клапан дымовой OKL-2D-30-1000x600", "шт", 2, 28500, 0, 5500)
    item("Решётка дымоудаления RDO-1000x600", "шт", 2, 8500, 0, 2000)
    item("Клапан нормально-закрытый OKL-2-30-800x500", "шт", 2, 24000, 0, 5000)
    item("Клапан дымовой OKL-2D-30-1200x500", "шт", 2, 32000, 0, 6000)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.fixture()
def ws():
    return Workspace.objects.create(name="WS-Demo", slug="ws-demo")


@pytest.fixture()
def user():
    return User.objects.create_user(username="demo-user", password="pass")


@pytest.mark.django_db
class TestDemoEstimate:
    def test_load_demo_creates_estimate_with_25_plus_items(self, ws, user):
        est = Estimate.objects.create(
            workspace=ws, name="Спецификация ОВ2 — демо",
            default_material_markup={"type": "percent", "value": 30},
            default_work_markup={"type": "percent", "value": 300},
            created_by=user,
        )
        xlsx = _make_demo_xlsx()
        result = import_estimate_xlsx(str(est.id), str(ws.id), xlsx)

        assert result.created >= 25
        assert result.errors == []
        sections = EstimateSection.objects.filter(estimate=est)
        assert sections.count() == 4
        items = EstimateItem.objects.filter(estimate=est, workspace_id=ws.id)
        assert items.count() >= 25

    def test_import_recalcs_totals(self, ws, user):
        est = Estimate.objects.create(
            workspace=ws, name="ОВ2 totals test",
            default_material_markup={"type": "percent", "value": 30},
            default_work_markup={"type": "percent", "value": 300},
            created_by=user,
        )
        xlsx = _make_demo_xlsx()
        import_estimate_xlsx(str(est.id), str(ws.id), xlsx)
        est.refresh_from_db()
        assert est.total_amount > 0
        assert est.total_equipment > 0
