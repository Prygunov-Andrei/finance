"""Тесты Excel export."""

import io

import pytest
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.estimate.excel.exporter import export_estimate_xlsx
from apps.estimate.models import Estimate, EstimateSection
from apps.estimate.services.estimate_service import EstimateService
from apps.workspace.models import Workspace

User = get_user_model()


@pytest.fixture()
def ws():
    return Workspace.objects.create(name="WS-Excel", slug="ws-excel")


@pytest.fixture()
def estimate(ws, user):
    return Estimate.objects.create(
        workspace=ws, name="Excel test",
        default_material_markup={"type": "percent", "value": 30},
        default_work_markup={"type": "percent", "value": 300},
        created_by=user,
    )


@pytest.fixture()
def user():
    return User.objects.create_user(username="excel-user", password="pass")


@pytest.fixture()
def section(estimate, ws):
    return EstimateSection.objects.create(
        estimate=estimate, workspace=ws, name="Вентиляция", sort_order=1
    )


@pytest.fixture()
def items(section, estimate, ws):
    names = ["Вентилятор", "Кабель UTP", "Датчик дыма"]
    created = []
    for i, name in enumerate(names):
        item = EstimateService.create_item(section, estimate, ws.id, {
            "name": name, "quantity": i + 1,
            "equipment_price": 100, "material_price": 200, "work_price": 50,
        })
        created.append(item)
    return created


@pytest.mark.django_db
class TestExcelExport:
    def test_export_creates_valid_xlsx(self, estimate, ws, section, items):
        output = export_estimate_xlsx(estimate.id, ws.id)
        assert isinstance(output, io.BytesIO)
        wb = load_workbook(output)
        assert "Смета" in wb.sheetnames
        assert "Агрегаты" in wb.sheetnames

    def test_sheet_contains_items_and_totals(self, estimate, ws, section, items):
        output = export_estimate_xlsx(estimate.id, ws.id)
        wb = load_workbook(output)
        ws_sheet = wb["Смета"]
        rows = list(ws_sheet.iter_rows(min_row=2, values_only=True))
        # Section header + 3 items + empty + totals = ≥5 rows
        names = [r[1] for r in rows if r[1] and r[1] not in ("ИТОГО",)]
        assert "Вентилятор" in names
        assert "Кабель UTP" in names
        assert "Датчик дыма" in names

    def test_row_id_hidden_column_present(self, estimate, ws, section, items):
        output = export_estimate_xlsx(estimate.id, ws.id)
        wb = load_workbook(output)
        ws_sheet = wb["Смета"]
        # TD-02: после добавления Модель/Производитель/Бренд/Система/
        # Примечание row_id переехал с колонки K (11) на P (16).
        assert ws_sheet.column_dimensions["P"].hidden is True
        assert ws_sheet.column_dimensions["Q"].hidden is True  # row_hash

    def test_tech_specs_columns_exported(self, estimate, ws, section):
        """TD-02 (#28): Модель/Производитель/Бренд/Система/Примечание
        должны экспортироваться из tech_specs."""
        EstimateService.create_item(section, estimate, ws.id, {
            "name": "Сплит-система 7 000 BTU",
            "quantity": 2,
            "material_price": 45000,
            "tech_specs": {
                "model_name": "RQ-71BV-A1",
                "manufacturer": "ООО «Фуджицу Россия»",
                "brand": "Fujitsu",
                "system": "К-01 Кондиционирование",
                "comments": "Установить до 2026-05-10",
                # прочие ключи tech_specs не экспортируются, но и не ломают ничего
                "power_kw": 2.2,
            },
        })

        output = export_estimate_xlsx(estimate.id, ws.id)
        wb = load_workbook(output)
        sheet = wb["Смета"]

        # Заголовки — колонки 3-6 (Модель/Производитель/Бренд/Система) + 15
        # (Примечание).
        headers = [sheet.cell(row=1, column=c).value for c in range(1, 18)]
        assert headers[2] == "Модель"
        assert headers[3] == "Производитель"
        assert headers[4] == "Бренд"
        assert headers[5] == "Система"
        assert headers[14] == "Примечание"

        # Строка с item — ищем по name в колонке 2.
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == "Сплит-система 7 000 BTU":
                assert row[2] == "RQ-71BV-A1"
                assert row[3] == "ООО «Фуджицу Россия»"
                assert row[4] == "Fujitsu"
                assert row[5] == "К-01 Кондиционирование"
                assert row[14] == "Установить до 2026-05-10"
                break
        else:
            pytest.fail("item row not found in exported sheet")

    def test_empty_tech_specs_exports_blank(self, estimate, ws, section, items):
        """Если tech_specs пустой — ячейки пустые (None/"" — openpyxl
        нормализует empty string → None при чтении). Главное — не падаем
        и не пишем мусорных значений, и importer потом корректно их игнорирует.
        """
        output = export_estimate_xlsx(estimate.id, ws.id)
        wb = load_workbook(output)
        sheet = wb["Смета"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == "Вентилятор":
                # Модель / Производитель / Бренд / Система / Примечание
                assert row[2] in (None, "")
                assert row[3] in (None, "")
                assert row[4] in (None, "")
                assert row[5] in (None, "")
                assert row[14] in (None, "")
                break
        else:
            pytest.fail("Вентилятор не найден в экспорте")

    def test_export_api_endpoint(self, estimate, ws, section, items, user):
        client = APIClient()
        client.force_authenticate(user=user)
        resp = client.get(
            f"/api/v1/estimates/{estimate.id}/export/xlsx/",
            HTTP_X_WORKSPACE_ID=str(ws.id),
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp["Content-Type"]
