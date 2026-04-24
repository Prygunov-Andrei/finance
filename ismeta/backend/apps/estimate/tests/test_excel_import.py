"""Тесты Excel import (E7)."""

import io
import uuid
from decimal import Decimal

import pytest
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook
from openpyxl.styles import Font
from rest_framework.test import APIClient

from apps.estimate.excel.importer import import_estimate_xlsx
from apps.estimate.models import Estimate, EstimateItem, EstimateSection
from apps.estimate.services.estimate_service import EstimateService
from apps.workspace.models import Workspace

User = get_user_model()
WS_HEADER = "HTTP_X_WORKSPACE_ID"


def _make_xlsx(rows, bold_rows=None):
    """Создать .xlsx в памяти. rows — list of lists. bold_rows — set of 0-based indices."""
    bold_rows = bold_rows or set()
    wb = Workbook()
    ws = wb.active
    ws.append(["Наименование", "Ед.изм.", "Кол-во", "Цена оборуд.", "Цена мат.", "Цена работ", "row_id"])
    for idx, row in enumerate(rows):
        ws.append(row)
        if idx in bold_rows:
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.fixture()
def ws():
    return Workspace.objects.create(name="WS-Import", slug="ws-import")


@pytest.fixture()
def user():
    return User.objects.create_user(username="import-user", password="pass")


@pytest.fixture()
def client(user):
    c = APIClient()
    c.force_authenticate(user=user)
    return c


@pytest.fixture()
def estimate(ws, user):
    return Estimate.objects.create(
        workspace=ws, name="Import test",
        default_material_markup={"type": "percent", "value": 30},
        default_work_markup={"type": "percent", "value": 300},
        created_by=user,
    )


@pytest.mark.django_db
class TestExcelImport:
    def test_import_new_items(self, estimate, ws):
        xlsx = _make_xlsx([
            ["Кабель UTP", "м", 100, 0, 150, 50, None],
            ["Вентилятор", "шт", 2, 85000, 0, 12000, None],
        ])
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 2
        assert result.updated == 0
        assert result.errors == []
        assert EstimateItem.objects.filter(estimate=estimate, workspace_id=ws.id).count() == 2

    def test_import_with_row_id_update(self, estimate, ws):
        section = EstimateSection.objects.create(
            estimate=estimate, workspace=ws, name="Тест", sort_order=1,
        )
        item = EstimateService.create_item(section, estimate, ws.id, {
            "name": "Кабель UTP", "unit": "м", "quantity": 50, "material_price": 100,
        })
        old_row_id = str(item.row_id)

        xlsx = _make_xlsx([
            ["Кабель UTP обновлённый", "м", 200, 0, 200, 50, old_row_id],
        ])
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.updated == 1
        assert result.created == 0

        updated = EstimateItem.all_objects.get(id=item.id)
        assert updated.name == "Кабель UTP обновлённый"
        assert updated.quantity == Decimal("200.0000")

    def test_import_with_sections(self, estimate, ws):
        xlsx = _make_xlsx(
            [
                ["Вентиляция", None, None, None, None, None, None],
                ["Воздуховод", "м.п.", 42, 0, 800, 200, None],
                ["Слаботочка", None, None, None, None, None, None],
                ["Кабель", "м", 100, 0, 150, 50, None],
            ],
            bold_rows={0, 2},
        )
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 2
        sections = EstimateSection.objects.filter(estimate=estimate)
        assert sections.count() == 2
        assert set(sections.values_list("name", flat=True)) == {"Вентиляция", "Слаботочка"}

    def test_import_empty_file(self, estimate, ws):
        xlsx = _make_xlsx([])
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 0
        assert result.updated == 0

    def test_import_invalid_format(self, estimate, ws):
        bad_file = io.BytesIO(b"not an xlsx")
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), bad_file)
        assert result.created == 0
        assert len(result.errors) > 0
        assert "Невалидный" in result.errors[0]

    def test_import_skips_bad_rows(self, estimate, ws):
        xlsx = _make_xlsx([
            [None, "м", 10, 0, 100, 50, None],  # пустое имя
            ["Кабель", "м", -5, 0, 100, 50, None],  # отрицательное кол-во
            ["Хороший кабель", "м", 10, 0, 100, 50, None],  # нормальный
        ])
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 1
        assert len(result.errors) == 2

    def test_import_recalcs_totals(self, estimate, ws):
        xlsx = _make_xlsx([
            ["Item 1", "шт", 10, 100, 200, 50, None],
            ["Item 2", "шт", 5, 50, 100, 30, None],
        ])
        import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        estimate.refresh_from_db()
        assert estimate.total_amount > 0

    # -----------------------------------------------------------------
    # TD-02 (DEV-BACKLOG #12): парсинг tech_specs-колонок UI-04
    # -----------------------------------------------------------------

    def _make_tech_specs_xlsx(self, rows, headers=None):
        """Вариант _make_xlsx с расширенными колонками UI-04."""
        headers = headers or [
            "Наименование", "Модель", "Производитель", "Бренд", "Система",
            "Ед.изм.", "Кол-во", "Цена оборуд.", "Цена мат.", "Цена работ",
            "Примечание", "row_id",
        ]
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_import_reads_model_manufacturer_brand(self, estimate, ws):
        """Модель/Производитель/Бренд → в tech_specs."""
        xlsx = self._make_tech_specs_xlsx([
            ["Сплит 7k", "RQ-71BV-A1", "ООО Фуджицу", "Fujitsu", "К-01",
             "шт", 2, 0, 45000, 0, "Установить до 10.05", None],
        ])
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 1
        assert result.errors == []

        item = EstimateItem.objects.get(estimate=estimate)
        assert item.name == "Сплит 7k"
        assert item.tech_specs.get("model_name") == "RQ-71BV-A1"
        assert item.tech_specs.get("manufacturer") == "ООО Фуджицу"
        assert item.tech_specs.get("brand") == "Fujitsu"
        assert item.tech_specs.get("system") == "К-01"
        assert item.tech_specs.get("comments") == "Установить до 10.05"

    def test_import_header_variations(self, estimate, ws):
        """Разные написания заголовков: Марка/Артикул/Изготовитель/Вендор/
        Notes/Контур — должны мапиться корректно."""
        xlsx = self._make_tech_specs_xlsx(
            [["Кабель ВВГнг-LS 3x2.5", "VV-3x2.5", "ЭКЗ", "ЭЛКАБ",
              "Э-01", "м", 100, 0, 120, 30, "bulk", None]],
            headers=["Наименование", "Марка", "Изготовитель", "Бренд",
                     "Контур", "Ед. изм.", "К-во", "Цена оборуд.",
                     "Цена мат.", "Цена работ", "Notes", "row_id"],
        )
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 1
        item = EstimateItem.objects.get(estimate=estimate)
        assert item.tech_specs.get("model_name") == "VV-3x2.5"
        assert item.tech_specs.get("manufacturer") == "ЭКЗ"
        assert item.tech_specs.get("brand") == "ЭЛКАБ"
        assert item.tech_specs.get("system") == "Э-01"
        assert item.tech_specs.get("comments") == "bulk"

    def test_import_empty_tech_specs_no_regression(self, estimate, ws):
        """Если колонки UI-04 пустые — tech_specs остаётся {}."""
        xlsx = self._make_tech_specs_xlsx([
            ["Только имя", "", "", "", "", "шт", 1, 0, 100, 0, "", None],
        ])
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 1
        item = EstimateItem.objects.get(estimate=estimate)
        # Пустые значения не попадают в tech_specs.
        assert item.tech_specs == {} or all(
            not item.tech_specs.get(k)
            for k in ("model_name", "brand", "manufacturer", "system", "comments")
        )

    def test_import_round_trip_preserves_existing_specs(self, estimate, ws):
        """Round-trip: при update по row_id tech_specs merge'ится, а не
        перезаписывается. Ключи, которых нет в Excel (power_kw и т.д.),
        должны остаться."""
        section = EstimateSection.objects.create(
            estimate=estimate, workspace=ws, name="Вентиляция", sort_order=1,
        )
        item = EstimateService.create_item(section, estimate, ws.id, {
            "name": "Вентилятор",
            "unit": "шт",
            "quantity": 1,
            "material_price": 50000,
            "tech_specs": {"power_kw": 2.2, "flow": "2600 м³/ч"},
        })
        old_row_id = str(item.row_id)

        xlsx = self._make_tech_specs_xlsx([
            ["Вентилятор обновлённый", "MOB-45", "MOB-LLC", "MOB", "В-01",
             "шт", 1, 0, 55000, 0, "поправили цену", old_row_id],
        ])
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.updated == 1
        assert result.created == 0

        updated = EstimateItem.all_objects.get(id=item.id)
        assert updated.name == "Вентилятор обновлённый"
        assert updated.tech_specs.get("model_name") == "MOB-45"
        assert updated.tech_specs.get("brand") == "MOB"
        assert updated.tech_specs.get("manufacturer") == "MOB-LLC"
        assert updated.tech_specs.get("system") == "В-01"
        assert updated.tech_specs.get("comments") == "поправили цену"
        # НЕ перезаписанные ключи сохранились.
        assert str(updated.tech_specs.get("power_kw")) == "2.2"
        assert updated.tech_specs.get("flow") == "2600 м³/ч"

    def test_import_cyrillic_model_with_hyphens(self, estimate, ws):
        """Edge-case: кириллица с дефисами «ВВГнг-LS», «ОВ-МД»."""
        xlsx = self._make_tech_specs_xlsx([
            ["Кабель", "ВВГнг-LS-3x2.5", "", "", "", "м", 50, 0, 110, 30, "", None],
        ])
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 1
        item = EstimateItem.objects.get(estimate=estimate)
        assert item.tech_specs.get("model_name") == "ВВГнг-LS-3x2.5"

    def test_import_no_tech_specs_columns_no_regression(self, estimate, ws):
        """Старый Excel без колонок UI-04 — всё ещё импортится корректно."""
        xlsx = _make_xlsx([
            ["Item 1", "шт", 10, 100, 200, 50, None],
        ])
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 1
        item = EstimateItem.objects.get(estimate=estimate)
        assert item.name == "Item 1"
        # tech_specs пустой — нет колонок не откуда взять.
        assert item.tech_specs == {}

    def test_multipart_upload_api(self, client, estimate, ws):
        xlsx = _make_xlsx([
            ["Кабель API", "м", 50, 0, 150, 50, None],
        ])
        resp = client.post(
            f"/api/v1/estimates/{estimate.id}/import/excel/",
            {"file": SimpleUploadedFile("test.xlsx", xlsx.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            format="multipart",
            **{WS_HEADER: str(ws.id)},
        )
        assert resp.status_code == 200
        assert resp.data["created"] == 1
