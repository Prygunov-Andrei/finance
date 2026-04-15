"""
Генерирует sample-estimate.xlsx из sample-estimate-data.json.

Запуск:
    pip install openpyxl
    python generate-sample-xlsx.py

Выход: sample-estimate.xlsx в этой же папке.

Формат листа соответствует specs/05-excel-schema.md.
Используется в тестах Excel-цикла и как наглядный пример для разработчиков.
"""
import hashlib
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.protection import SheetProtection

HERE = Path(__file__).parent
DATA_PATH = HERE / "sample-estimate-data.json"
OUT_PATH = HERE / "sample-estimate.xlsx"

COLUMNS = [
    ("row_id", 0),                 # A — скрытый
    ("row_hash", 0),               # B — скрытый
    ("№", 5),                      # C
    ("Раздел", 30),                # D
    ("Подраздел", 25),             # E
    ("Наименование", 50),          # F
    ("Модель/Артикул", 20),        # G
    ("Бренд", 15),                 # H
    ("Ед.", 8),                    # I
    ("Количество", 15),            # J
    ("Цена закуп. материал", 18),  # K
    ("Цена закуп. работа", 18),    # L
    ("Наценка материал, %", 16),   # M
    ("Наценка работа, %", 16),     # N
    ("Продажная материал", 18),    # O
    ("Продажная работа", 18),      # P
    ("Итого материал", 18),        # Q
    ("Итого работа", 18),          # R
    ("Итого", 18),                 # S
    ("Примечание", 30),            # T
]


def _row_hash(row: dict) -> str:
    payload = {
        "name": row["name"],
        "model_name": row.get("model_name", ""),
        "unit": row["unit"],
        "quantity": str(row["quantity"]),
        "material_unit_price": str(row.get("material_unit_price", 0)),
        "work_unit_price": str(row.get("work_unit_price", 0)),
    }
    return hashlib.sha256(json.dumps(payload, sort_keys=True, ensure_ascii=False).encode()).hexdigest()[:16]


def generate() -> None:
    with open(DATA_PATH, encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()

    # Лист 1: Смета
    ws = wb.active
    if ws is None:
        raise RuntimeError("Could not create active sheet")
    ws.title = "Смета"

    # Заголовки
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width if width > 0 else 0.1

    # Скрываем колонки A и B (row_id, row_hash)
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].hidden = True

    # Пробегаемся по разделам и строкам
    row_no = 2
    item_counter = 0
    for section in data["sections"]:
        # Строка раздела
        ws.cell(row=row_no, column=4, value=section["name"])
        section_cell = ws.cell(row=row_no, column=4)
        section_cell.font = Font(bold=True)
        section_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        row_no += 1

        for item in section["items"]:
            item_counter += 1
            ws.cell(row=row_no, column=1, value=item["row_id"])
            ws.cell(row=row_no, column=2, value=_row_hash(item))
            ws.cell(row=row_no, column=3, value=item_counter)
            ws.cell(row=row_no, column=6, value=item["name"])
            ws.cell(row=row_no, column=7, value=item.get("model_name", ""))
            ws.cell(row=row_no, column=8, value=item.get("brand", ""))
            ws.cell(row=row_no, column=9, value=item["unit"])
            ws.cell(row=row_no, column=10, value=item["quantity"])
            ws.cell(row=row_no, column=11, value=item.get("material_unit_price", 0))
            ws.cell(row=row_no, column=12, value=item.get("work_unit_price", 0))
            # M и N пустые (наценки на уровне сметы по умолчанию)
            # O (продажная материал) = K*(1 + M/100) — тут 30%
            mat_markup = data["defaults"]["material_markup_percent"]
            work_markup = data["defaults"]["work_markup_percent"]
            ws.cell(row=row_no, column=15, value=f"=K{row_no}*(1+{mat_markup}/100)")
            ws.cell(row=row_no, column=16, value=f"=L{row_no}*(1+{work_markup}/100)")
            ws.cell(row=row_no, column=17, value=f"=J{row_no}*O{row_no}")
            ws.cell(row=row_no, column=18, value=f"=J{row_no}*P{row_no}")
            ws.cell(row=row_no, column=19, value=f"=Q{row_no}+R{row_no}")
            row_no += 1

    # Frozen header
    ws.freeze_panes = "A2"

    # Защита листа (без пароля, чтобы не потерять)
    ws.protection.sheet = False  # оставляем редактируемым; в prod можно включить с паролем

    # Лист 2: _ismeta_meta
    meta_ws = wb.create_sheet("_ismeta_meta")
    meta_ws.sheet_state = "hidden"
    meta_row = 1
    for key, value in data["meta"].items():
        meta_ws.cell(row=meta_row, column=1, value=key)
        meta_ws.cell(row=meta_row, column=2, value=str(value))
        meta_row += 1

    # Лист 3: Инструкция
    inst_ws = wb.create_sheet("Инструкция")
    instructions = [
        "Инструкция по работе с Excel-циклом ISMeta",
        "",
        "Этот файл выгружен из ISMeta для массовой правки в Excel.",
        "",
        "Что можно менять в листе «Смета»:",
        "  — Наименование, модель, бренд, ед., количество;",
        "  — Цены закупочные (K, L);",
        "  — Наценки на уровне строки (M, N) — оставляй пустыми, если используются дефолтные;",
        "  — Примечание (T);",
        "  — Добавлять новые строки в существующий раздел;",
        "  — Удалять строки (будут помечены как удалённые при импорте).",
        "",
        "Что НЕЛЬЗЯ:",
        "  — Трогать скрытые столбцы A (row_id) и B (row_hash);",
        "  — Менять схему файла (удалять столбцы, переставлять местами);",
        "  — Редактировать лист «_ismeta_meta».",
        "",
        "Сохранять:",
        "  — Только в формате .xlsx (не .csv, не .xls).",
        "",
        "После правки — в ISMeta нажми «Импортировать обновления»,",
        "система покажет diff-предпросмотр перед применением.",
    ]
    for idx, line in enumerate(instructions, start=1):
        cell = inst_ws.cell(row=idx, column=1, value=line)
        if idx == 1:
            cell.font = Font(bold=True, size=14)

    inst_ws.column_dimensions["A"].width = 80
    inst_ws.protection = SheetProtection(sheet=True)

    wb.save(OUT_PATH)
    print(f"✅ Создан {OUT_PATH} ({OUT_PATH.stat().st_size} байт)")


if __name__ == "__main__":
    generate()
