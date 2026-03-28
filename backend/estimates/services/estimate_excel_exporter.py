"""
Excel-экспорт сметы.

Два режима:
1. export() / export_public() — публичный портал, 3 секции (основное/аналоги/уточнение).
2. export_with_column_config() — экспорт с учётом column_config (формулы, custom-столбцы).

При export_public=True — применяет наценку из PublicPricingConfig.
"""
import logging
from decimal import Decimal
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from estimates.models import Estimate, EstimateItem, EstimateSection

logger = logging.getLogger(__name__)


def get_sale_price(purchase_price: Decimal, category=None) -> Decimal:
    """Рассчитывает продажную цену: закупочная × (1 + наценка/100).

    Наценка берётся из PublicPricingConfig (каскад: категория → родитель → default).
    """
    if not purchase_price or purchase_price <= 0:
        return Decimal('0')

    from api_public.models import PublicPricingConfig
    markup = PublicPricingConfig.get_markup(category)
    return (purchase_price * (Decimal('1') + markup / Decimal('100'))).quantize(Decimal('0.01'))


class EstimateExcelExporter:
    """Генерация Excel-файла сметы.

    Использование:
        exporter = EstimateExcelExporter(estimate)
        buffer = exporter.export()           # закупочные цены
        buffer = exporter.export_public()    # с наценкой для клиента
    """

    def __init__(self, estimate: Estimate):
        self.estimate = estimate

    def export(self, apply_markup: bool = False) -> BytesIO:
        """Генерирует Excel-файл сметы.

        Args:
            apply_markup: True = применить наценку из PublicPricingConfig.

        Returns:
            BytesIO с .xlsx файлом.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Смета — Материалы и Работы'

        # Стили
        title_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12, color='1F4E79')
        header_font = Font(bold=True, size=10)
        bold_font = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        wrap = Alignment(wrap_text=True, vertical='top')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        num_fmt = '#,##0.00'
        yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        red_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')

        # Заголовок
        ws.merge_cells('A1:I1')
        ws['A1'] = f'СМЕТА'
        ws['A1'].font = title_font
        ws['A1'].alignment = center

        ws['A2'] = f'Проект: {self.estimate.name}'
        ws['A3'] = f'Дата: {self.estimate.created_at.strftime("%d.%m.%Y")}'

        # Получаем все позиции
        items = EstimateItem.objects.filter(
            estimate=self.estimate,
        ).select_related(
            'section', 'product', 'product__category',
        ).order_by('section__sort_order', 'sort_order', 'item_number')

        # Классификация позиций
        exact_items = []
        analog_items = []
        unknown_items = []

        for item in items:
            if not item.product_id and item.material_unit_price == 0:
                unknown_items.append(item)
            elif item.is_analog:
                analog_items.append(item)
            else:
                exact_items.append(item)

        row = 5
        total_materials = Decimal('0')
        total_works = Decimal('0')

        # Секция 1: Основное оборудование
        if exact_items:
            row = self._write_main_section(
                ws, row, 'РАЗДЕЛ 1: ОСНОВНОЕ ОБОРУДОВАНИЕ',
                exact_items, apply_markup,
                section_font, header_font, bold_font, center, wrap,
                thin_border, num_fmt,
            )
            subtotals = self._calc_subtotals(exact_items, apply_markup)
            total_materials += subtotals['materials']
            total_works += subtotals['works']
            row += 1

        # Секция 2: Аналоги
        if analog_items:
            row = self._write_analog_section(
                ws, row, 'РАЗДЕЛ 2: АНАЛОГИ',
                analog_items, apply_markup,
                section_font, header_font, bold_font, center, wrap,
                thin_border, num_fmt, yellow_fill,
            )
            subtotals = self._calc_subtotals(analog_items, apply_markup)
            total_materials += subtotals['materials']
            total_works += subtotals['works']
            row += 1

        # Секция 3: Требует уточнения
        if unknown_items:
            row = self._write_unknown_section(
                ws, row, 'РАЗДЕЛ 3: ТРЕБУЕТ УТОЧНЕНИЯ',
                unknown_items,
                section_font, header_font, center, wrap,
                thin_border, red_fill,
            )
            row += 1

        # Итоги
        row += 1
        ws.cell(row=row, column=1, value='Итого материалы:').font = bold_font
        ws.cell(row=row, column=8, value=float(total_materials)).number_format = num_fmt
        ws.cell(row=row, column=8).font = bold_font
        row += 1

        ws.cell(row=row, column=1, value='Итого работы:').font = bold_font
        ws.cell(row=row, column=9, value=float(total_works)).number_format = num_fmt
        ws.cell(row=row, column=9).font = bold_font
        row += 1

        grand_total = total_materials + total_works
        ws.cell(row=row, column=1, value='ИТОГО (без НДС):').font = bold_font
        ws.cell(row=row, column=9, value=float(grand_total)).number_format = num_fmt
        ws.cell(row=row, column=9).font = bold_font

        if self.estimate.with_vat:
            row += 1
            vat_rate = Decimal(str(self.estimate.vat_rate))
            vat_amount = (grand_total * vat_rate / Decimal('100')).quantize(Decimal('0.01'))
            ws.cell(row=row, column=1, value=f'НДС {self.estimate.vat_rate}%:').font = bold_font
            ws.cell(row=row, column=9, value=float(vat_amount)).number_format = num_fmt
            row += 1
            ws.cell(row=row, column=1, value='ИТОГО С НДС:').font = Font(bold=True, size=12)
            ws.cell(row=row, column=9, value=float(grand_total + vat_amount)).number_format = num_fmt
            ws.cell(row=row, column=9).font = Font(bold=True, size=12)

        row += 2
        ws.cell(row=row, column=1, value='* Позиции раздела "Требует уточнения" не включены в итоговую сумму')
        ws.cell(row=row + 1, column=1, value='* Цены актуальны на дату составления сметы')

        # Ширина колонок
        col_widths = [5, 30, 15, 6, 8, 14, 14, 14, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def export_public(self) -> BytesIO:
        """Экспорт с наценкой для публичного портала."""
        return self.export(apply_markup=True)

    def _get_price(self, item: EstimateItem, apply_markup: bool) -> Decimal:
        """Получает цену материала (с наценкой или без)."""
        price = item.material_unit_price or Decimal('0')
        if apply_markup and price > 0:
            category = item.product.category if item.product_id else None
            return get_sale_price(price, category)
        return price

    def _write_main_section(self, ws, row, title, items, apply_markup,
                            section_font, header_font, bold_font, center, wrap,
                            thin_border, num_fmt):
        """Записывает секцию основного оборудования."""
        # Заголовок секции
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.cell(row=row, column=1, value=title).font = section_font
        row += 1

        # Шапка таблицы
        headers = ['#', 'Наименование', 'Модель', 'Ед.', 'Кол.',
                   'Материал, ₽', 'Работа, ₽', 'Мат. итого', 'Раб. итого']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border
        row += 1

        for i, item in enumerate(items, 1):
            mat_price = self._get_price(item, apply_markup)
            work_price = item.work_unit_price or Decimal('0')
            qty = item.quantity or Decimal('0')

            values = [
                i, item.name, item.model_name or '', item.unit,
                float(qty), float(mat_price), float(work_price),
                float(mat_price * qty), float(work_price * qty),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                if isinstance(val, float):
                    cell.number_format = num_fmt
            row += 1

        return row

    def _write_analog_section(self, ws, row, title, items, apply_markup,
                              section_font, header_font, bold_font, center, wrap,
                              thin_border, num_fmt, fill):
        """Записывает секцию аналогов."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.cell(row=row, column=1, value=title).font = section_font
        row += 1

        headers = ['#', 'Запрошено', 'Предложено', 'Обоснование', 'Ед.', 'Кол.',
                   'Материал, ₽', 'Работа, ₽', 'Итого, ₽']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border
        row += 1

        for i, item in enumerate(items, 1):
            mat_price = self._get_price(item, apply_markup)
            work_price = item.work_unit_price or Decimal('0')
            qty = item.quantity or Decimal('0')
            line_total = (mat_price + work_price) * qty

            values = [
                i,
                item.original_name or item.name,
                item.name,
                item.analog_reason or '',
                item.unit,
                float(qty),
                float(mat_price * qty),
                float(work_price * qty),
                float(line_total),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.fill = fill
                if isinstance(val, float):
                    cell.number_format = num_fmt
            row += 1

        return row

    def _write_unknown_section(self, ws, row, title, items,
                               section_font, header_font, center, wrap,
                               thin_border, fill):
        """Записывает секцию 'Требует уточнения'."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.cell(row=row, column=1, value=title).font = section_font
        row += 1

        headers = ['#', 'Наименование', 'Модель', 'Ед.', 'Кол.', '', '', '', 'Примечание']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border
        row += 1

        for i, item in enumerate(items, 1):
            values = [
                i, item.name, item.model_name or '', item.unit,
                float(item.quantity or 0), '', '', '',
                'Нет в каталоге, цена по запросу',
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.fill = fill
            row += 1

        return row

    def _calc_subtotals(self, items, apply_markup: bool) -> dict:
        """Считает итоги по секции."""
        materials = Decimal('0')
        works = Decimal('0')
        for item in items:
            qty = item.quantity or Decimal('0')
            mat_price = self._get_price(item, apply_markup)
            work_price = item.work_unit_price or Decimal('0')
            materials += mat_price * qty
            works += work_price * qty
        return {'materials': materials, 'works': works}

    # ── Экспорт с column_config ─────────────────────────────────────

    # Колонки, скрываемые в external-режиме (закупочные цены, наценки)
    _INTERNAL_ONLY_KEYS = {
        'material_unit_price', 'work_unit_price',
        'material_total', 'work_total', 'line_total',
        'material_purchase_total', 'work_purchase_total',
        'effective_material_markup_percent', 'effective_work_markup_percent',
    }

    def export_with_column_config(self, mode: str = 'internal') -> BytesIO:
        """Генерирует Excel-файл сметы с учётом column_config.

        Args:
            mode: 'internal' — все колонки (закупка + наценка + продажа).
                  'external' — только продажные цены, без закупок и наценок.

        Returns:
            BytesIO с .xlsx файлом.
        """
        from estimates.column_defaults import DEFAULT_COLUMN_CONFIG
        from estimates.formula_engine import compute_all_formulas

        estimate = self.estimate
        config = estimate.column_config or DEFAULT_COLUMN_CONFIG
        visible_cols = [c for c in config if c.get('visible', True)]

        if mode == 'external':
            visible_cols = [c for c in visible_cols if c['key'] not in self._INTERNAL_ONLY_KEYS]

        items = list(EstimateItem.objects.filter(
            estimate=estimate,
        ).select_related('section', 'estimate').order_by(
            'section__sort_order', 'sort_order', 'item_number',
        ))

        sections = EstimateSection.objects.filter(
            estimate=estimate,
        ).order_by('sort_order')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Смета'

        bold = Font(bold=True)
        header_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=11, color='1F4E79')
        center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        num_fmt = '#,##0.00'

        # Title
        num_cols = len(visible_cols)
        last_col_letter = openpyxl.utils.get_column_letter(max(num_cols, 1))
        ws.merge_cells(f'A1:{last_col_letter}1')
        ws['A1'] = f'Смета №{estimate.number} — {estimate.name}'
        ws['A1'].font = header_font
        ws['A1'].alignment = center

        # Column headers (row 3)
        for col_idx, col_def in enumerate(visible_cols, 1):
            label = col_def.get('label', col_def['key'])
            # В external-режиме "продажные" колонки отображаются как просто "Цена мат." и т.д.
            if mode == 'external':
                label = label.replace('Продажа ', 'Цена ').replace('Итого продажа ', 'Итого ')
            cell = ws.cell(row=3, column=col_idx, value=label)
            cell.font = bold
            cell.alignment = center
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(
                col_def.get('width', 100) / 8, 8,
            )

        row_num = 4
        agg_sums = {c['key']: Decimal('0') for c in visible_cols if c.get('aggregatable')}

        for section in sections:
            # Section header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_cols)
            cell = ws.cell(row=row_num, column=1, value=section.name)
            cell.font = section_font
            row_num += 1

            section_items = [i for i in items if i.section_id == section.id]
            for item in section_items:
                builtin_values = {
                    'item_number': Decimal(str(item.item_number or 0)),
                    'quantity': item.quantity or Decimal('0'),
                    'material_unit_price': item.material_unit_price or Decimal('0'),
                    'work_unit_price': item.work_unit_price or Decimal('0'),
                    'material_total': item.material_total or Decimal('0'),
                    'work_total': item.work_total or Decimal('0'),
                    'line_total': item.line_total or Decimal('0'),
                    # Новые builtin-поля для наценок
                    'material_sale_unit_price': item.material_sale_unit_price or Decimal('0'),
                    'work_sale_unit_price': item.work_sale_unit_price or Decimal('0'),
                    'material_purchase_total': item.material_purchase_total or Decimal('0'),
                    'work_purchase_total': item.work_purchase_total or Decimal('0'),
                    'material_sale_total': item.material_sale_total or Decimal('0'),
                    'work_sale_total': item.work_sale_total or Decimal('0'),
                    'effective_material_markup_percent': item.effective_material_markup_percent or Decimal('0'),
                    'effective_work_markup_percent': item.effective_work_markup_percent or Decimal('0'),
                }
                custom_data = item.custom_data or {}
                computed = compute_all_formulas(config, builtin_values, custom_data)

                for col_idx, col_def in enumerate(visible_cols, 1):
                    key = col_def['key']
                    col_type = col_def.get('type', 'builtin')
                    value = None

                    if col_type == 'builtin':
                        field = col_def.get('builtin_field', key)
                        # Сначала проверяем builtin_values (включает новые поля)
                        if key in builtin_values:
                            value = builtin_values[key]
                        else:
                            value = getattr(item, field, None)
                    elif col_type == 'formula':
                        value = computed.get(key)
                    elif col_type.startswith('custom_'):
                        value = custom_data.get(key, '')

                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.border = thin_border

                    if isinstance(value, Decimal):
                        cell.value = float(value)
                        cell.number_format = num_fmt
                    elif value is not None:
                        try:
                            cell.value = float(value)
                            cell.number_format = num_fmt
                        except (ValueError, TypeError):
                            cell.value = str(value) if value else ''
                    else:
                        cell.value = ''

                    # Accumulate aggregatables
                    if key in agg_sums and value is not None:
                        try:
                            agg_sums[key] += Decimal(str(value))
                        except Exception:
                            pass

                row_num += 1

        # Footer totals
        if agg_sums:
            row_num += 1
            for col_idx, col_def in enumerate(visible_cols, 1):
                key = col_def['key']
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = bold
                cell.border = thin_border
                if key in agg_sums:
                    cell.value = float(agg_sums[key])
                    cell.number_format = num_fmt
                elif col_idx == 1:
                    cell.value = 'ИТОГО'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
