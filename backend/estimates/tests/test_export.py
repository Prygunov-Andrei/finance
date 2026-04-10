"""
Тесты Excel-экспорта сметы.
~8 тест-кейсов.
"""
from decimal import Decimal
from io import BytesIO
from django.test import TestCase, RequestFactory
from django.contrib.auth.models import User
import openpyxl

from estimates.models import Estimate, EstimateSection, EstimateItem
from estimates.views import EstimateViewSet
from estimates.column_defaults import DEFAULT_COLUMN_CONFIG


def _make_estimate_with_items(user, column_config=None):
    """Helper: смета + раздел + несколько строк."""
    from accounting.models import LegalEntity, TaxSystem
    from objects.models import Object as BuildObject

    obj, _ = BuildObject.objects.get_or_create(name='Тест-экспорт', defaults={'address': 'тест'})
    ts, _ = TaxSystem.objects.get_or_create(name='УСН', defaults={'code': 'usn', 'has_vat': False})
    le, _ = LegalEntity.objects.get_or_create(
        short_name='ТестООО',
        defaults={'name': 'ООО Тест', 'inn': '1234567890', 'tax_system': ts},
    )
    estimate = Estimate.objects.create(
        object=obj, legal_entity=le, name='Тест-смета',
        created_by=user, column_config=column_config or [],
    )
    section = EstimateSection.objects.create(
        estimate=estimate, name='Раздел 1', sort_order=1,
    )
    EstimateItem.objects.create(
        estimate=estimate, section=section, name='Кабель ВВГнг',
        quantity=Decimal('100'), material_unit_price=Decimal('50'),
        work_unit_price=Decimal('20'), sort_order=1, item_number=1,
    )
    EstimateItem.objects.create(
        estimate=estimate, section=section, name='Автомат ABB',
        quantity=Decimal('5'), material_unit_price=Decimal('800'),
        work_unit_price=Decimal('200'), sort_order=2, item_number=2,
    )
    return estimate, section


class TestEstimateExport(TestCase):

    def setUp(self):
        self.user = User.objects.create_user('exporter', password='test')
        self.factory = RequestFactory()

    def _export(self, estimate):
        request = self.factory.get(f'/api/v1/estimates/{estimate.id}/export/')
        request.user = self.user
        view = EstimateViewSet.as_view({'get': 'export'})
        response = view(request, pk=estimate.pk)
        return response

    def test_export_default_columns(self):
        """Экспорт с дефолтными столбцами — xlsx файл с 10 столбцами."""
        estimate, _ = _make_estimate_with_items(self.user)
        response = self._export(estimate)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        # Row 3 = headers, should have 10 columns for default config
        headers = [ws.cell(row=3, column=i).value for i in range(1, 11)]
        self.assertIn('Наименование', headers)
        self.assertIn('Итого закупка', headers)

    def test_export_content_disposition(self):
        estimate, _ = _make_estimate_with_items(self.user)
        response = self._export(estimate)
        cd = response['Content-Disposition']
        # Content-Disposition может быть RFC 2047-encoded (base64) при кириллице
        from email.header import decode_header
        decoded_parts = decode_header(cd)
        decoded_cd = ''.join(
            part.decode(enc or 'utf-8') if isinstance(part, bytes) else part
            for part, enc in decoded_parts
        )
        self.assertIn('attachment', decoded_cd)
        self.assertIn('.xlsx', decoded_cd)

    def test_export_custom_columns(self):
        """Экспорт с кастомным column_config — столбцы соответствуют."""
        config = [
            {'key': 'name', 'type': 'builtin', 'builtin_field': 'name', 'label': 'Наименование',
             'width': 250, 'editable': True, 'visible': True, 'formula': None,
             'decimal_places': None, 'aggregatable': False, 'options': None},
            {'key': 'quantity', 'type': 'builtin', 'builtin_field': 'quantity', 'label': 'Кол-во',
             'width': 80, 'editable': True, 'visible': True, 'formula': None,
             'decimal_places': 3, 'aggregatable': False, 'options': None},
            {'key': 'note', 'type': 'custom_text', 'label': 'Примечание',
             'width': 150, 'editable': True, 'visible': True, 'formula': None,
             'decimal_places': None, 'aggregatable': False, 'options': None},
        ]
        estimate, _ = _make_estimate_with_items(self.user, column_config=config)
        response = self._export(estimate)
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [ws.cell(row=3, column=i).value for i in range(1, 4)]
        self.assertEqual(headers, ['Наименование', 'Кол-во', 'Примечание'])

    def test_hidden_columns_excluded(self):
        """Скрытые столбцы не попадают в Excel."""
        config = [
            {'key': 'name', 'type': 'builtin', 'builtin_field': 'name', 'label': 'Наименование',
             'width': 250, 'editable': True, 'visible': True, 'formula': None,
             'decimal_places': None, 'aggregatable': False, 'options': None},
            {'key': 'model_name', 'type': 'builtin', 'builtin_field': 'model_name', 'label': 'Модель',
             'width': 150, 'editable': True, 'visible': False, 'formula': None,
             'decimal_places': None, 'aggregatable': False, 'options': None},
        ]
        estimate, _ = _make_estimate_with_items(self.user, column_config=config)
        response = self._export(estimate)
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [ws.cell(row=3, column=i).value for i in range(1, 3)]
        self.assertEqual(headers[0], 'Наименование')
        self.assertNotEqual(headers[1] if len(headers) > 1 else None, 'Модель')

    def test_section_headers_in_export(self):
        """Секции отображаются как merged rows."""
        estimate, _ = _make_estimate_with_items(self.user)
        response = self._export(estimate)
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        # Row 4 should be section name "Раздел 1"
        self.assertEqual(ws.cell(row=4, column=1).value, 'Раздел 1')

    def test_formula_columns_computed(self):
        """Формульные столбцы вычислены корректно."""
        config = [
            {'key': 'quantity', 'type': 'builtin', 'builtin_field': 'quantity', 'label': 'Кол-во',
             'width': 80, 'editable': True, 'visible': True, 'formula': None,
             'decimal_places': 3, 'aggregatable': False, 'options': None},
            {'key': 'material_unit_price', 'type': 'builtin', 'builtin_field': 'material_unit_price',
             'label': 'Цена мат.', 'width': 100, 'editable': True, 'visible': True,
             'formula': None, 'decimal_places': 2, 'aggregatable': False, 'options': None},
            {'key': 'mat_total', 'type': 'formula', 'formula': 'quantity * material_unit_price',
             'label': 'Итого мат.', 'width': 120, 'editable': False, 'visible': True,
             'decimal_places': 2, 'aggregatable': True, 'options': None},
        ]
        estimate, _ = _make_estimate_with_items(self.user, column_config=config)
        response = self._export(estimate)
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        # Row 5 = first item (row 4 is section header)
        # Col 3 = mat_total = 100 * 50 = 5000
        val = ws.cell(row=5, column=3).value
        self.assertAlmostEqual(float(val), 5000.0, places=2)

    def test_empty_estimate_export(self):
        """Пустая смета (0 items) — корректный xlsx."""
        from accounting.models import LegalEntity, TaxSystem
        from objects.models import Object as BuildObject

        obj, _ = BuildObject.objects.get_or_create(name='Пустой', defaults={'address': 'тест'})
        ts, _ = TaxSystem.objects.get_or_create(name='УСН', defaults={'code': 'usn', 'has_vat': False})
        le, _ = LegalEntity.objects.get_or_create(
            short_name='ТестООО',
            defaults={'name': 'ООО Тест', 'inn': '1234567890', 'tax_system': ts},
        )
        estimate = Estimate.objects.create(
            object=obj, legal_entity=le, name='Пустая смета', created_by=self.user,
        )
        response = self._export(estimate)
        self.assertEqual(response.status_code, 200)

    def test_aggregatable_totals(self):
        """Итоги (aggregatable) в футере — суммы корректны."""
        config = [
            {'key': 'name', 'type': 'builtin', 'builtin_field': 'name', 'label': 'Наименование',
             'width': 250, 'editable': True, 'visible': True, 'formula': None,
             'decimal_places': None, 'aggregatable': False, 'options': None},
            {'key': 'material_total', 'type': 'builtin', 'builtin_field': 'material_total',
             'label': 'Итого мат.', 'width': 110, 'editable': False, 'visible': True,
             'formula': None, 'decimal_places': 2, 'aggregatable': True, 'options': None},
        ]
        estimate, _ = _make_estimate_with_items(self.user, column_config=config)
        response = self._export(estimate)
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active

        # Find footer row (after items)
        # Items: row 5 = item 1, row 6 = item 2, row 8 = footer (row 7 = blank)
        # material_total: 100*50=5000, 5*800=4000 => total = 9000
        footer_val = None
        for row in range(7, 12):
            cell = ws.cell(row=row, column=2)
            if cell.value is not None:
                try:
                    footer_val = float(cell.value)
                except (ValueError, TypeError):
                    continue
                if footer_val > 8000:
                    break

        self.assertIsNotNone(footer_val)
        self.assertAlmostEqual(footer_val, 9000.0, places=0)
