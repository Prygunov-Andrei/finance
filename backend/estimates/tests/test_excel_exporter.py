"""Тесты EstimateExcelExporter — Заход 3."""
import pytest
from decimal import Decimal

import openpyxl

from estimates.models import (
    Estimate, EstimateSection, EstimateSubsection, EstimateItem,
)
from estimates.services.estimate_excel_exporter import (
    EstimateExcelExporter, get_sale_price,
)
from api_public.models import PublicPricingConfig
from api_public.tests.factories import PublicPricingConfigFactory


@pytest.fixture
def tax_system(db):
    from accounting.models import TaxSystem
    ts, _ = TaxSystem.objects.get_or_create(code='osno', defaults={'name': 'ОСНО'})
    return ts


@pytest.fixture
def legal_entity(db, tax_system):
    from accounting.models import LegalEntity
    return LegalEntity.objects.create(
        name='ООО Тест', short_name='Тест', tax_system=tax_system,
    )


@pytest.fixture
def portal_object(db):
    from objects.models import Object
    return Object.objects.create(name='Тест портал')


@pytest.fixture
def system_user(db):
    from django.contrib.auth.models import User
    return User.objects.create_user(username='exporter_test')


@pytest.fixture
def estimate_with_items(db, portal_object, legal_entity, system_user):
    """Смета с позициями разных типов: exact, analog, unknown."""
    from catalog.models import Category, Product

    cat = Category.objects.create(name='Вентиляция', code='vent')
    product = Product.objects.create(
        name='Вентилятор ВКК-160', category=cat,
        normalized_name='вентилятор вкк 160',
    )

    estimate = Estimate.objects.create(
        name='Тест экспорт', object=portal_object,
        legal_entity=legal_entity, created_by=system_user,
    )

    section = EstimateSection.objects.create(
        estimate=estimate, name='ОВ', sort_order=0,
    )
    sub = EstimateSubsection.objects.create(
        section=section, name='Оборудование', sort_order=0,
    )

    # Exact match
    EstimateItem.objects.create(
        estimate=estimate, section=section, subsection=sub,
        name='Вентилятор ВКК-160', unit='шт', quantity=3,
        material_unit_price=Decimal('10000'), work_unit_price=Decimal('2000'),
        product=product,
    )

    # Analog
    EstimateItem.objects.create(
        estimate=estimate, section=section, subsection=sub,
        name='Вентилятор ВКК-200', unit='шт', quantity=2,
        material_unit_price=Decimal('15000'), work_unit_price=Decimal('2500'),
        product=product, is_analog=True,
        analog_reason='Другой типоразмер', original_name='Вентилятор ВКК-250',
    )

    # Unknown (no product, no price)
    EstimateItem.objects.create(
        estimate=estimate, section=section, subsection=sub,
        name='Контроллер Carel pCO5', model_name='pCO5', unit='шт', quantity=1,
        material_unit_price=Decimal('0'), work_unit_price=Decimal('0'),
    )

    return estimate


@pytest.fixture
def empty_estimate(db, portal_object, legal_entity, system_user):
    """Пустая смета (без позиций)."""
    return Estimate.objects.create(
        name='Пустая смета', object=portal_object,
        legal_entity=legal_entity, created_by=system_user,
    )


def _load_wb(buffer):
    """Загрузить workbook из BytesIO."""
    return openpyxl.load_workbook(buffer)


class TestGetSalePrice:
    """Тесты get_sale_price()."""

    def test_zero_price(self, db):
        """Нулевая цена → 0."""
        assert get_sale_price(Decimal('0')) == Decimal('0')

    def test_negative_price(self, db):
        """Отрицательная цена → 0."""
        assert get_sale_price(Decimal('-100')) == Decimal('0')

    def test_default_markup(self, db):
        """Без конфигов → 30% по умолчанию."""
        result = get_sale_price(Decimal('10000'))
        assert result == Decimal('13000.00')

    def test_custom_default_markup(self, db):
        """С default PublicPricingConfig → используется его наценка."""
        PublicPricingConfigFactory(is_default=True, markup_percent='25.00')
        result = get_sale_price(Decimal('10000'))
        assert result == Decimal('12500.00')

    def test_category_markup(self, db):
        """Наценка по категории."""
        from catalog.models import Category
        cat = Category.objects.create(name='Трубы', code='pipes')
        PublicPricingConfigFactory(is_default=True, markup_percent='30.00')
        PublicPricingConfigFactory(category=cat, markup_percent='15.00')
        result = get_sale_price(Decimal('10000'), category=cat)
        assert result == Decimal('11500.00')


class TestEstimateExcelExporter:
    """Тесты экспортера."""

    def test_export_produces_valid_xlsx(self, estimate_with_items):
        """Экспорт создаёт валидный .xlsx файл."""
        exporter = EstimateExcelExporter(estimate_with_items)
        buffer = exporter.export()
        wb = _load_wb(buffer)
        assert wb.active is not None
        assert wb.active.title == 'Смета — Материалы и Работы'

    def test_export_has_title(self, estimate_with_items):
        """Первая строка — заголовок 'СМЕТА'."""
        exporter = EstimateExcelExporter(estimate_with_items)
        buffer = exporter.export()
        wb = _load_wb(buffer)
        ws = wb.active
        assert ws['A1'].value == 'СМЕТА'

    def test_export_without_markup(self, estimate_with_items):
        """export(apply_markup=False) — закупочные цены."""
        exporter = EstimateExcelExporter(estimate_with_items)
        buffer = exporter.export(apply_markup=False)
        wb = _load_wb(buffer)
        ws = wb.active

        # Ищем строку с "Вентилятор ВКК-160" (exact) — цена 10000
        found = False
        for row in ws.iter_rows(min_row=5, max_col=9, values_only=True):
            if row[1] == 'Вентилятор ВКК-160' and not found:
                # Колонка 6 (индекс 5) — цена материала за ед.
                assert row[5] == 10000.0
                found = True
        assert found, 'Не найдена позиция "Вентилятор ВКК-160"'

    def test_export_with_markup(self, estimate_with_items):
        """export_public() — цены с наценкой 30% (default)."""
        exporter = EstimateExcelExporter(estimate_with_items)
        buffer = exporter.export_public()
        wb = _load_wb(buffer)
        ws = wb.active

        # Ищем строку с "Вентилятор ВКК-160" (exact) — цена 10000 * 1.3 = 13000
        found = False
        for row in ws.iter_rows(min_row=5, max_col=9, values_only=True):
            if row[1] == 'Вентилятор ВКК-160' and not found:
                assert row[5] == 13000.0
                found = True
        assert found, 'Не найдена позиция "Вентилятор ВКК-160"'

    def test_three_sections_present(self, estimate_with_items):
        """Все 3 секции присутствуют в Excel."""
        exporter = EstimateExcelExporter(estimate_with_items)
        buffer = exporter.export()
        wb = _load_wb(buffer)
        ws = wb.active

        section_titles = []
        for row in ws.iter_rows(max_col=1, values_only=True):
            val = row[0]
            if isinstance(val, str) and val.startswith('РАЗДЕЛ'):
                section_titles.append(val)

        assert len(section_titles) == 3
        assert 'ОСНОВНОЕ' in section_titles[0]
        assert 'АНАЛОГИ' in section_titles[1]
        assert 'УТОЧНЕНИЯ' in section_titles[2]

    def test_analog_section_has_reason(self, estimate_with_items):
        """Секция аналогов содержит обоснование."""
        exporter = EstimateExcelExporter(estimate_with_items)
        buffer = exporter.export()
        wb = _load_wb(buffer)
        ws = wb.active

        found_reason = False
        for row in ws.iter_rows(min_row=5, max_col=9, values_only=True):
            if row and 'Другой типоразмер' in str(row):
                found_reason = True
        assert found_reason, 'Обоснование аналога не найдено'

    def test_unknown_section_no_prices(self, estimate_with_items):
        """Секция 'Требует уточнения' — без цен."""
        exporter = EstimateExcelExporter(estimate_with_items)
        buffer = exporter.export()
        wb = _load_wb(buffer)
        ws = wb.active

        found_unknown = False
        for row in ws.iter_rows(min_row=5, max_col=9, values_only=True):
            if row and row[1] == 'Контроллер Carel pCO5':
                found_unknown = True
                # Примечание в последней колонке
                assert 'по запросу' in str(row[8]).lower() or 'каталоге' in str(row[8]).lower()
        assert found_unknown

    def test_empty_estimate(self, empty_estimate):
        """Пустая смета → Excel с заголовками, без данных."""
        exporter = EstimateExcelExporter(empty_estimate)
        buffer = exporter.export()
        wb = _load_wb(buffer)
        ws = wb.active
        assert ws['A1'].value == 'СМЕТА'

    def test_totals_exclude_unknown(self, estimate_with_items):
        """Итого НЕ включает позиции 'Требует уточнения'."""
        exporter = EstimateExcelExporter(estimate_with_items)
        buffer = exporter.export(apply_markup=False)
        wb = _load_wb(buffer)
        ws = wb.active

        # Exact: 10000*3 = 30000 материалы, 2000*3 = 6000 работы
        # Analog: 15000*2 = 30000 материалы, 2500*2 = 5000 работы
        # Unknown: 0 (не включены)
        # Итого материалы: 60000, работы: 11000

        found_total = False
        for row in ws.iter_rows(max_col=9, values_only=True):
            if row and row[0] == 'Итого материалы:':
                assert row[7] == 60000.0
                found_total = True
        assert found_total

    def test_markup_does_not_affect_model(self, estimate_with_items):
        """Наценка применяется только в Excel, не меняет EstimateItem."""
        exporter = EstimateExcelExporter(estimate_with_items)
        exporter.export_public()

        # Проверяем что в БД цена не изменилась
        item = EstimateItem.objects.filter(
            estimate=estimate_with_items, name='Вентилятор ВКК-160',
            is_analog=False,
        ).first()
        assert item.material_unit_price == Decimal('10000')
