"""Тесты сервисов import_template и model_import."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from ac_catalog.models import ACModel, ModelRawValue
from ac_catalog.services import generate_import_template_xlsx
from ac_catalog.services.model_import import (
    find_existing_models_in_file,
    import_models_from_file,
)
from ac_methodology.models import Criterion, MethodologyCriterion
from ac_methodology.tests.factories import ActiveMethodologyVersionFactory


def _make_active_methodology_with_binary_criterion(code: str = "erv"):
    mv = ActiveMethodologyVersionFactory(version=f"imp-{code}")
    criterion = Criterion.objects.create(
        code=code, name_ru="ЭРВ", value_type=Criterion.ValueType.BINARY,
    )
    MethodologyCriterion.objects.create(
        methodology=mv, criterion=criterion,
        scoring_type=MethodologyCriterion.ScoringType.BINARY,
        weight=100, display_order=1,
    )
    return mv, criterion


def _build_xlsx_with_one_model(tmp_path: Path, criterion_code: str) -> Path:
    """Создаёт .xlsx с заголовками (FIXED + criterion_code) и одной строкой."""
    from ac_catalog.services.import_template import FIXED_COLUMNS

    headers = FIXED_COLUMNS + [criterion_code]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    row = [
        "TestBrand",  # brand
        "model-x",    # model
        "outer-x",    # outer_unit
        "S",          # series
        "2500",       # nominal_capacity
        "30000",      # price
        "Сплит-система",  # equipment_type
        "ru",         # region
        "",           # youtube
        "",           # rutube
        "",           # vk
        "",           # compressor_model
        "да",         # criterion value
    ]
    ws.append(row)
    path = tmp_path / "import.xlsx"
    wb.save(str(path))
    return path


# ── generate_import_template_xlsx ─────────────────────────────────────


@pytest.mark.django_db
def test_generate_import_template_no_active_methodology_raises():
    with pytest.raises(ValueError, match="методики"):
        generate_import_template_xlsx()


@pytest.mark.django_db
def test_generate_import_template_returns_xlsx_with_headers():
    _make_active_methodology_with_binary_criterion(code="erv")
    body, fname = generate_import_template_xlsx()
    assert fname.endswith(".xlsx")

    wb = openpyxl.load_workbook(BytesIO(body))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    # FIXED_COLUMNS + criterion code
    assert "brand" in headers
    assert "model" in headers
    assert "erv" in headers

    # Лист "Критерии" описывает текущую методику
    assert "Критерии" in wb.sheetnames
    ws2 = wb["Критерии"]
    crit_codes = [ws2.cell(row=r, column=1).value for r in range(2, ws2.max_row + 1)]
    assert "erv" in crit_codes


# ── import_models_from_file ───────────────────────────────────────────


@pytest.mark.django_db
def test_import_creates_model_with_brand_and_raw_value(tmp_path):
    _make_active_methodology_with_binary_criterion(code="erv")
    path = _build_xlsx_with_one_model(tmp_path, "erv")

    imported, errors = import_models_from_file(path)

    assert imported == 1
    assert errors == []  # никаких предупреждений на чистом импорте
    ac = ACModel.objects.get(brand__name="TestBrand", inner_unit="MODEL-X")
    assert ac.publish_status == ACModel.PublishStatus.DRAFT
    assert ac.nominal_capacity == 2500.0
    rv = ModelRawValue.objects.get(model=ac, criterion__code="erv")
    assert rv.raw_value == "да"


@pytest.mark.django_db
def test_import_publish_flag_publishes_model(tmp_path):
    _make_active_methodology_with_binary_criterion(code="erv")
    path = _build_xlsx_with_one_model(tmp_path, "erv")

    import_models_from_file(path, publish=True)
    ac = ACModel.objects.get(inner_unit="MODEL-X")
    assert ac.publish_status == ACModel.PublishStatus.PUBLISHED


@pytest.mark.django_db
def test_import_existing_model_does_not_duplicate(tmp_path):
    _make_active_methodology_with_binary_criterion(code="erv")
    path = _build_xlsx_with_one_model(tmp_path, "erv")

    # Первый импорт создаёт
    n1, _ = import_models_from_file(path)
    assert n1 == 1
    assert ACModel.objects.filter(brand__name="TestBrand", inner_unit="MODEL-X").count() == 1

    # Второй импорт — модель уже есть, дубликата быть не должно;
    # imported == 0, но в errors предупреждение «модель уже существует».
    n2, errors2 = import_models_from_file(path)
    assert n2 == 0
    assert ACModel.objects.filter(brand__name="TestBrand", inner_unit="MODEL-X").count() == 1
    assert any("уже существует" in e for e in errors2)


@pytest.mark.django_db
def test_find_existing_models_in_file_returns_label(tmp_path):
    _make_active_methodology_with_binary_criterion(code="erv")
    path = _build_xlsx_with_one_model(tmp_path, "erv")
    # До импорта — пусто
    assert find_existing_models_in_file(path) == []
    # После — модель появляется
    import_models_from_file(path)
    found = find_existing_models_in_file(path)
    # Имя модели нормализуется в верхний регистр (_normalize_model_name).
    assert found == ["TestBrand MODEL-X"]
