from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

from core.version_mixin import VersioningMixin
from .models import (
    WorkerGrade, WorkSection, WorkerGradeSkills,
    WorkItem, PriceList, PriceListAgreement, PriceListItem
)
from .serializers import (
    WorkerGradeSerializer, WorkSectionSerializer, WorkerGradeSkillsSerializer,
    WorkItemSerializer, WorkItemListSerializer,
    PriceListSerializer, PriceListListSerializer, PriceListCreateSerializer,
    PriceListAgreementSerializer, PriceListItemSerializer,
    AddRemoveItemsSerializer
)


class WorkerGradeViewSet(viewsets.ModelViewSet):
    """ViewSet для разрядов рабочих"""
    
    queryset = WorkerGrade.objects.prefetch_related('skills', 'skills__section')
    serializer_class = WorkerGradeSerializer
    http_method_names = ['get', 'post', 'patch', 'head', 'options']
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_active']


class WorkSectionViewSet(viewsets.ModelViewSet):
    """ViewSet для разделов работ"""
    
    queryset = WorkSection.objects.select_related('parent').prefetch_related('children')
    serializer_class = WorkSectionSerializer
    http_method_names = ['get', 'post', 'patch', 'head', 'options']
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['is_active', 'parent']
    search_fields = ['code', 'name']

    def get_queryset(self):
        queryset = super().get_queryset()
        # Если запрошен параметр tree=true, возвращаем только корневые разделы
        if self.request.query_params.get('tree') == 'true':
            queryset = queryset.filter(parent__isnull=True)
        return queryset
    
    @action(detail=False, methods=['get'])
    def tree(self, request):
        """
        Возвращает иерархическое дерево разделов работ.
        Оптимизировано: загружает всё дерево одним запросом.
        """
        from collections import defaultdict
        
        # Загружаем ВСЕ активные разделы ОДНИМ запросом
        all_sections = list(
            WorkSection.objects.filter(is_active=True)
            .order_by('sort_order', 'code', 'name')
            .values('id', 'code', 'name', 'parent_id')
        )
        
        # Строим дерево в памяти
        sections_by_parent = defaultdict(list)
        for section in all_sections:
            sections_by_parent[section['parent_id']].append(section)
        
        def build_tree(parent_id):
            result = []
            for section in sections_by_parent.get(parent_id, []):
                result.append({
                    'id': section['id'],
                    'code': section['code'],
                    'name': section['name'],
                    'children': build_tree(section['id'])
                })
            return result
        
        return Response(build_tree(None))


class WorkerGradeSkillsViewSet(viewsets.ModelViewSet):
    """ViewSet для навыков разрядов"""
    
    queryset = WorkerGradeSkills.objects.select_related('grade', 'section')
    serializer_class = WorkerGradeSkillsSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['grade', 'section']


class WorkItemViewSet(VersioningMixin, viewsets.ModelViewSet):
    """ViewSet для работ (с поддержкой версионирования через VersioningMixin)"""
    
    queryset = WorkItem.objects.select_related('section', 'grade', 'parent_version')
    http_method_names = ['get', 'post', 'patch', 'head', 'options']
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['section', 'grade', 'is_current']
    search_fields = ['article', 'name']
    version_list_serializer_class = WorkItemListSerializer

    def get_serializer_class(self):
        if self.action == 'list':
            return WorkItemListSerializer
        return WorkItemSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        # По умолчанию показываем только актуальные версии
        if 'is_current' not in self.request.query_params:
            queryset = queryset.filter(is_current=True)
        return queryset

    def partial_update(self, request, *args, **kwargs):
        """При обновлении создаём новую версию работы"""
        instance = self.get_object()
        
        # Создаём новую версию
        new_version = instance.create_new_version()
        
        # Обновляем новую версию переданными данными
        serializer = self.get_serializer(new_version, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        
        # Применяем изменения к новой версии (кроме артикула и версионных полей)
        for attr, value in serializer.validated_data.items():
            if attr not in ['article', 'version_number', 'is_current', 'parent_version']:
                setattr(new_version, attr, value)
        new_version.save()
        
        return Response(WorkItemSerializer(new_version).data)
    
    # Метод versions() наследуется от VersioningMixin


class PriceListViewSet(viewsets.ModelViewSet):
    """ViewSet для прайс-листов"""
    
    queryset = PriceList.objects.prefetch_related(
        'items',
        'items__work_item',
        'items__work_item__section',
        'items__work_item__grade',
        'agreements',
        'agreements__counterparty'
    ).select_related('parent_version')
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['status']
    search_fields = ['number', 'name']

    def get_serializer_class(self):
        if self.action == 'list':
            return PriceListListSerializer
        if self.action == 'create':
            return PriceListCreateSerializer
        return PriceListSerializer

    def get_queryset(self):
        from django.db.models import Count, Q, Sum
        from django.db.models.functions import Coalesce
        from decimal import Decimal
        
        queryset = super().get_queryset()
        
        # Фильтрация по дате
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        
        # Добавляем annotate для list и retrieve view (оптимизация N+1)
        if self.action in ['list', 'retrieve']:
            queryset = queryset.annotate(
                annotated_items_count=Count('items', filter=Q(items__is_included=True)),
                annotated_agreements_count=Count('agreements')
            )
        
        return queryset

    @action(detail=True, methods=['post'], url_path='create-version')
    def create_version(self, request, pk=None):
        """Создать новую версию прайс-листа"""
        price_list = self.get_object()
        new_version = price_list.create_new_version()
        serializer = PriceListSerializer(new_version)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='add-items')
    def add_items(self, request, pk=None):
        """Добавить работы в прайс-лист"""
        price_list = self.get_object()
        serializer = AddRemoveItemsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        work_item_ids = serializer.validated_data['work_item_ids']
        work_items = WorkItem.objects.filter(id__in=work_item_ids, is_current=True)
        work_items_dict = {wi.id: wi for wi in work_items}
        
        # Находим существующие элементы
        existing_items = PriceListItem.objects.filter(
            price_list=price_list,
            work_item__in=work_items
        ).select_related('work_item')
        existing_work_item_ids = {item.work_item_id for item in existing_items}
        
        added = []
        
        # Обновляем неактивные элементы через bulk_update
        items_to_update = [
            item for item in existing_items 
            if not item.is_included
        ]
        if items_to_update:
            for item in items_to_update:
                item.is_included = True
                added.append(item.work_item_id)
            PriceListItem.objects.bulk_update(items_to_update, ['is_included'])
        
        # Создаём новые элементы через bulk_create
        new_work_items = [
            work_items_dict[wid] for wid in work_item_ids 
            if wid in work_items_dict and wid not in existing_work_item_ids
        ]
        if new_work_items:
            new_items = [
                PriceListItem(price_list=price_list, work_item=wi, is_included=True)
                for wi in new_work_items
            ]
            PriceListItem.objects.bulk_create(new_items)
            added.extend([wi.id for wi in new_work_items])
        
        return Response({
            'added': added,
            'count': len(added)
        })

    @action(detail=True, methods=['post'], url_path='remove-items')
    def remove_items(self, request, pk=None):
        """Удалить работы из прайс-листа"""
        price_list = self.get_object()
        serializer = AddRemoveItemsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        work_item_ids = serializer.validated_data['work_item_ids']
        deleted_count = PriceListItem.objects.filter(
            price_list=price_list,
            work_item_id__in=work_item_ids
        ).delete()[0]
        
        return Response({
            'removed': work_item_ids,
            'count': deleted_count
        })

    @action(detail=True, methods=['get'])
    def export(self, request, pk=None):
        """Экспорт прайс-листа в Excel"""
        price_list = self.get_object()
        
        # Создаём Excel-файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Прайс-лист"
        
        # Стили
        bold_font = Font(bold=True)
        header_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        ws.merge_cells('A1:I1')
        ws['A1'] = f"Прайс-лист №{price_list.number} от {price_list.date.strftime('%d.%m.%Y')}"
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align
        
        # Название (если есть)
        if price_list.name:
            ws.merge_cells('A2:I2')
            ws['A2'] = price_list.name
            ws['A2'].alignment = center_align
        
        # Ставки по разрядам
        row = 4
        ws[f'A{row}'] = "Ставки по разрядам:"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        for grade_num in range(1, 6):
            rate = price_list.get_rate_for_grade(grade_num)
            ws[f'A{row}'] = f"Разряд {grade_num}:"
            ws[f'B{row}'] = f"{rate} руб/ч"
            row += 1
        
        # Пустая строка
        row += 1
        
        # Заголовки таблицы работ
        ws[f'A{row}'] = "Работы:"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        headers = ['Артикул', 'Раздел', 'Наименование', 'Ед.изм.', 'Часы', 'Разряд', 'Коэфф.', 'Стоимость', 'Комментарий']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
        
        # Устанавливаем ширину столбцов
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 30  # Комментарий
        
        row += 1
        
        # Данные работ
        for item in price_list.items.filter(is_included=True).select_related(
            'work_item', 'work_item__section', 'work_item__grade'
        ):
            work = item.work_item
            # Используем эффективный разряд (может быть дробным)
            effective_grade = float(item.effective_grade)
            values = [
                work.article,
                work.section.code,
                work.name,
                work.unit,
                float(item.effective_hours),
                effective_grade,  # Дробный разряд
                float(item.effective_coefficient),
                float(item.calculated_cost),
                work.comment if work.comment else ''  # Комментарий
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col in [5, 6, 7, 8]:  # Часы, Разряд, Коэфф., Стоимость
                    cell.alignment = center_align
                elif col == 9:  # Комментарий - выравнивание по левому краю
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            row += 1
        
        # Итого
        total = sum(
            item.calculated_cost
            for item in price_list.items.filter(is_included=True)
        )
        row += 1
        ws[f'G{row}'] = "ИТОГО:"
        ws[f'G{row}'].font = bold_font
        ws[f'H{row}'] = float(total)
        ws[f'H{row}'].font = bold_font
        # Объединяем ячейки для строки "Итого" (столбец комментария)
        ws.merge_cells(f'I{row}:I{row}')
        
        # Сохраняем в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Формируем ответ
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"pricelist_{price_list.number}_{price_list.date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class PriceListItemViewSet(viewsets.ModelViewSet):
    """ViewSet для позиций прайс-листа"""
    
    queryset = PriceListItem.objects.select_related(
        'price_list', 'work_item', 'work_item__section', 'work_item__grade'
    )
    serializer_class = PriceListItemSerializer
    http_method_names = ['get', 'post', 'patch', 'delete', 'head', 'options']
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['price_list', 'is_included']


class PriceListAgreementViewSet(viewsets.ModelViewSet):
    """ViewSet для согласований прайс-листов"""
    
    queryset = PriceListAgreement.objects.select_related('price_list', 'counterparty')
    serializer_class = PriceListAgreementSerializer
    http_method_names = ['get', 'post', 'delete', 'head', 'options']
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['price_list', 'counterparty']
