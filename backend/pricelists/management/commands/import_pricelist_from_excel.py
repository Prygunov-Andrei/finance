from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from decimal import Decimal, InvalidOperation
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pricelists.models import WorkSection, WorkItem, WorkerGrade
from datetime import date


class Command(BaseCommand):
    help = """
    Импортирует работы из Excel файла в базу данных.
    
    Структура Excel файла:
    - Каждая закладка (лист) = Раздел работ
    - Строки с жёлтой заливкой = Подразделы (создаются как WorkSection с parent)
    - Столбцы:
      A - Артикул работы
      B - Наименование работы
      C - Единица измерения
      D - Состав работ (описание)
      E - Время (часы)
      F - Разрядность (может быть дробной, например 2.5, 3.65)
      G - Комментарий (опционально)
    
    Пример использования:
      python manage.py import_pricelist_from_excel path/to/file.xlsx
      python manage.py import_pricelist_from_excel path/to/file.xlsx --dry-run
      python manage.py import_pricelist_from_excel path/to/file.xlsx --skip-header
    """

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='Путь к Excel файлу (.xlsx)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Пробный запуск без сохранения в БД'
        )
        parser.add_argument(
            '--skip-header',
            action='store_true',
            help='Пропустить первую строку (заголовки)'
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        dry_run = options['dry_run']

        if dry_run:
            self.stdout.write(self.style.WARNING('РЕЖИМ ПРОБНОГО ЗАПУСКА - изменения не будут сохранены'))

        try:
            wb = load_workbook(excel_file, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'Файл не найден: {excel_file}')
        except Exception as e:
            raise CommandError(f'Ошибка при открытии файла: {e}')

        # Проверяем наличие разрядов в БД
        grades = WorkerGrade.objects.filter(is_active=True)
        if not grades.exists():
            raise CommandError('В базе данных нет активных разрядов. Запустите команду populate_pricelists сначала.')

        # Маппинг единиц измерения
        unit_mapping = self.get_unit_mapping()

        stats = {
            'sections_created': 0,
            'subsections_created': 0,
            'work_items_created': 0,
            'work_items_updated': 0,
            'errors': [],
            'warnings': []
        }

        with transaction.atomic():
            # Обрабатываем каждый лист (раздел)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                self.stdout.write(f'\nОбработка раздела: {sheet_name}')

                # Создаём или находим раздел
                # Генерируем код раздела из названия листа
                section_code = self.generate_section_code(sheet_name)
                section, created = WorkSection.objects.get_or_create(
                    code=section_code,
                    defaults={
                        'name': sheet_name[:200],  # Ограничение длины
                        'sort_order': 0
                    }
                )
                if created:
                    stats['sections_created'] += 1
                    self.stdout.write(f'  ✓ Создан раздел: {section.name}')

                current_subsection = None  # Текущий подраздел (для жёлтых строк)

                # Определяем начальную строку
                # Если skip_header, начинаем со строки 2, но проверяем строку 1 на подраздел
                start_row = 2 if options['skip_header'] else 1
                
                # Если skip_header, проверяем первую строку на подраздел
                if options['skip_header']:
                    first_row = ws[1]
                    article = first_row[0].value if len(first_row) > 0 else None
                    name = first_row[1].value if len(first_row) > 1 else None
                    
                    if name and (not article or (isinstance(article, str) and article.strip() == '')):
                        # Проверяем, что остальные столбцы пустые
                        has_work_data = False
                        for col_idx in [2, 3, 4, 5, 6]:
                            if col_idx < len(first_row) and first_row[col_idx].value:
                                val = str(first_row[col_idx].value).strip()
                                if val and val.lower() not in ['', '0', '0.0', '0.00']:
                                    has_work_data = True
                                    break
                        
                        if not has_work_data:
                            # Создаём подраздел из первой строки
                            subsection_name = str(name).strip()
                            base_code = f"{section.code}-SUB1"[:50]
                            subsection_code = base_code
                            counter = 1
                            while WorkSection.objects.filter(code=subsection_code).exists():
                                subsection_code = f"{base_code}{counter}"[:50]
                                counter += 1
                            
                            current_subsection, created = WorkSection.objects.get_or_create(
                                code=subsection_code,
                                defaults={
                                    'name': subsection_name[:200],
                                    'parent': section,
                                    'sort_order': 1
                                }
                            )
                            if created:
                                stats['subsections_created'] += 1
                                self.stdout.write(f'  ✓ Создан подраздел: {current_subsection.name}')
                
                # Парсим строки
                for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=False), start=start_row):
                    # Пропускаем пустые строки
                    if not row[0].value and not row[1].value:
                        continue

                    # Определяем подраздел по правилу:
                    # Подраздел = столбец A (артикул) пустой ИЛИ содержит специальный маркер (SUB, РАЗДЕЛ, ПОДРАЗДЕЛ)
                    # И столбец B (название) заполнен
                    # И столбцы C, D, E, F, G (единица, состав, часы, разряд, комментарий) пустые или почти пустые
                    article_value = row[0].value
                    name_value = row[1].value
                    
                    # Проверяем, является ли строка подразделом
                    is_subsection = False
                    subsection_name = None
                    
                    if name_value:  # Название должно быть заполнено
                        name_str = str(name_value).strip()
                        
                        # Вариант 1: Артикул пустой или None
                        if not article_value or (isinstance(article_value, str) and article_value.strip() == ''):
                            # Проверяем, что остальные столбцы (C, D, E, F, G) пустые или почти пустые
                            # Это означает, что это подраздел, а не работа
                            has_work_data = False
                            for col_idx in [2, 3, 4, 5, 6]:  # C, D, E, F, G
                                if col_idx < len(row) and row[col_idx].value:
                                    cell_val = str(row[col_idx].value).strip()
                                    if cell_val and cell_val.lower() not in ['', '0', '0.0', '0.00']:
                                        has_work_data = True
                                        break
                            
                            if not has_work_data:
                                is_subsection = True
                                subsection_name = name_str
                        
                        # Вариант 2: Артикул содержит специальный маркер
                        elif article_value:
                            article_str = str(article_value).strip().upper()
                            markers = ['SUB', 'РАЗДЕЛ', 'ПОДРАЗДЕЛ', 'SECTION', 'SUBSECTION']
                            if any(marker in article_str for marker in markers):
                                is_subsection = True
                                subsection_name = name_str
                    
                    if is_subsection and subsection_name:
                        # Создаём подраздел
                        # Создаём уникальный код подраздела
                        base_code = f"{section.code}-SUB{row_idx}"[:50]
                        subsection_code = base_code
                        counter = 1
                        while WorkSection.objects.filter(code=subsection_code).exists():
                            subsection_code = f"{base_code}{counter}"[:50]
                            counter += 1
                        
                        current_subsection, created = WorkSection.objects.get_or_create(
                            code=subsection_code,
                            defaults={
                                'name': subsection_name[:200],  # Ограничение длины
                                'parent': section,
                                'sort_order': row_idx
                            }
                        )
                        if created:
                            stats['subsections_created'] += 1
                            self.stdout.write(f'  ✓ Создан подраздел: {current_subsection.name}')
                        continue

                    # Обычная строка - работа
                    try:
                        work_item = self.parse_work_item_row(
                            row, 
                            section if not current_subsection else current_subsection,
                            unit_mapping,
                            stats
                        )
                        if work_item:
                            if not dry_run:
                                work_item.save()
                                stats['work_items_created' if work_item.version_number == 1 else 'work_items_updated'] += 1
                            self.stdout.write(f'  ✓ {"Создана" if work_item.version_number == 1 else "Обновлена"} работа: {work_item.article} - {work_item.name}')
                    except Exception as e:
                        error_msg = f'Строка {row_idx}: {str(e)}'
                        stats['errors'].append(error_msg)
                        self.stdout.write(self.style.ERROR(f'  ✗ Ошибка: {error_msg}'))

            if dry_run:
                # Откатываем транзакцию в режиме пробного запуска
                transaction.set_rollback(True)

        # Выводим статистику
        self.stdout.write('\n' + '='*50)
        self.stdout.write('СТАТИСТИКА ИМПОРТА:')
        self.stdout.write(f'  Разделов создано: {stats["sections_created"]}')
        self.stdout.write(f'  Подразделов создано: {stats["subsections_created"]}')
        self.stdout.write(f'  Работ создано: {stats["work_items_created"]}')
        self.stdout.write(f'  Работ обновлено: {stats["work_items_updated"]}')
        self.stdout.write(f'  Ошибок: {len(stats["errors"])}')
        self.stdout.write(f'  Предупреждений: {len(stats["warnings"])}')

        if stats['warnings']:
            self.stdout.write('\nПРЕДУПРЕЖДЕНИЯ:')
            for warning in stats['warnings']:
                self.stdout.write(self.style.WARNING(f'  - {warning}'))

        if stats['errors']:
            self.stdout.write('\nОШИБКИ:')
            for error in stats['errors']:
                self.stdout.write(self.style.ERROR(f'  - {error}'))

        if dry_run:
            self.stdout.write(self.style.WARNING('\nЭто был пробный запуск. Для реального импорта уберите флаг --dry-run'))
        else:
            self.stdout.write(self.style.SUCCESS('\nИмпорт завершён успешно!'))

    def is_yellow_row(self, cell):
        """Проверяет, является ли ячейка жёлтой (подраздел)"""
        if not cell or not cell.fill:
            return False
        
        try:
            fill = cell.fill
            if not isinstance(fill, PatternFill):
                return False
            
            # Проверяем цвет заливки через RGB
            if fill.start_color:
                # Проверяем RGB значение
                if hasattr(fill.start_color, 'rgb') and fill.start_color.rgb:
                    rgb = str(fill.start_color.rgb).upper().replace('#', '')
                    
                    # Обрабатываем ARGB формат (8 символов: AARRGGBB)
                    # Например: FFFFFF00 = жёлтый (FF FF 00 в RGB)
                    if len(rgb) == 8:
                        # Пропускаем альфа-канал (первые 2 символа), берём RGB
                        r = rgb[2:4]
                        g = rgb[4:6]
                        b = rgb[6:8]
                    elif len(rgb) >= 6:
                        # Обычный RGB формат (6 символов)
                        r = rgb[0:2]
                        g = rgb[2:4]
                        b = rgb[4:6]
                    else:
                        return False
                    
                    # Жёлтый = высокий R и G, низкий B
                    try:
                        r_val = int(r, 16) if r.isalnum() else 0
                        g_val = int(g, 16) if g.isalnum() else 0
                        b_val = int(b, 16) if b.isalnum() else 0
                        # Проверяем, что это жёлтый оттенок
                        if r_val > 200 and g_val > 200 and b_val < 100:
                            return True
                    except (ValueError, TypeError):
                        pass
                    
                    # Проверяем точное совпадение с известными жёлтыми цветами
                    yellow_patterns = [
                        'FFEB3B', 'FFFE00', 'FFFF00', 'FFD700', 'FFC107',  # Яркие жёлтые
                        'FFF9C4', 'FFF59D', 'FFF176', 'FFEE58',  # Светло-жёлтые
                        'FFFFFF00',  # ARGB формат жёлтого
                    ]
                    for pattern in yellow_patterns:
                        if pattern in rgb or rgb.endswith(pattern):
                            return True
                
                # Проверяем через индекс цвета (для стандартных цветов Excel)
                if hasattr(fill.start_color, 'index') and fill.start_color.index:
                    yellow_indices = ['FFFF00', 'FFEB3B', '40']  # Жёлтый в Excel
                    if str(fill.start_color.index) in yellow_indices:
                        return True
                
                # Проверяем через theme
                if hasattr(fill.start_color, 'theme') and fill.start_color.theme:
                    # Theme 4 в Excel обычно жёлтый
                    if fill.start_color.theme == 4:
                        return True
        except Exception:
            # В случае ошибки считаем, что строка не жёлтая
            pass
        
        return False

    def generate_section_code(self, name):
        """Генерирует код раздела из названия"""
        # Убираем спецсимволы, оставляем только буквы, цифры и пробелы
        clean_name = re.sub(r'[^\w\s]', '', str(name))
        # Берем первые буквы каждого слова
        words = clean_name.split()
        if words:
            code = ''.join([w[0].upper() for w in words if w])[:50]
        else:
            # Если нет слов, берем первые символы
            code = clean_name.upper()[:50]
        
        # Если код пустой, используем имя
        if not code:
            code = clean_name.upper()[:50]
        
        # Убеждаемся, что код уникален
        base_code = code
        counter = 1
        while WorkSection.objects.filter(code=code).exists():
            code = f"{base_code}{counter}"[:50]
            counter += 1
        
        return code

    def get_unit_mapping(self):
        """Маппинг произвольных единиц измерения к стандартным"""
        return {
            'шт': 'шт',
            'шт.': 'шт',  # С точкой
            'штука': 'шт',
            'штуки': 'шт',
            'м.п.': 'м.п.',
            'м.п': 'м.п.',
            'м п': 'м.п.',
            'метр погонный': 'м.п.',
            'м²': 'м²',
            'м2': 'м²',
            'м.кв.': 'м²',  # С точкой
            'кв.м': 'м²',
            'кв м': 'м²',
            'квадратный метр': 'м²',
            'м³': 'м³',
            'м3': 'м³',
            'куб.м': 'м³',
            'куб м': 'м³',
            'кубический метр': 'м³',
            'компл': 'компл',
            'компл.': 'компл',  # С точкой
            'комплект': 'компл',
            'комп': 'компл',
            'ед': 'ед',
            'единица': 'ед',
            'единицы': 'ед',
            'ч': 'ч',
            'час': 'ч',
            'часы': 'ч',
            'часов': 'ч',
            'кг': 'кг',
            'кг.': 'кг',  # С точкой
            'килограмм': 'кг',
            'килограммы': 'кг',
            'т': 'т',
            'тонна': 'т',
            'тонны': 'т',
            'точка': 'точка',  # Отдельная единица
        }

    def parse_work_item_row(self, row, section, unit_mapping, stats):
        """
        Парсит строку работы из Excel
        Столбцы: A - артикул, B - наименование, C - единица, D - состав, E - время, F - разрядность, G - комментарий
        """
        # A - Артикул (может быть числом или строкой)
        article_value = row[0].value
        if article_value is None:
            raise ValueError('Артикул не указан')
        # Преобразуем в строку (если число - убираем лишние нули)
        if isinstance(article_value, (int, float)):
            article = str(int(article_value)) if article_value == int(article_value) else str(article_value)
        else:
            article = str(article_value).strip()
        if not article:
            raise ValueError('Артикул не указан')

        # B - Наименование
        name = str(row[1].value).strip() if row[1].value else None
        if not name:
            raise ValueError('Наименование не указано')

        # C - Единица измерения
        unit_raw = str(row[2].value).strip().lower() if row[2].value else 'ед'
        unit = unit_mapping.get(unit_raw, 'ед')  # По умолчанию 'ед'
        if unit_raw not in unit_mapping:
            stats['warnings'].append(f'Неизвестная единица измерения "{unit_raw}" для работы {article}, использовано "ед"')

        # D - Состав работ
        composition = str(row[3].value).strip() if row[3].value else ''

        # E - Время (часы, опционально, по умолчанию 0)
        hours_value = row[4].value
        if hours_value is None:
            hours = Decimal('0')  # Если не указано, используем 0
        else:
            try:
                hours = Decimal(str(hours_value))
                if hours < 0:
                    raise ValueError('Время не может быть отрицательным')
            except (InvalidOperation, ValueError) as e:
                raise ValueError(f'Некорректное значение времени: {hours_value}')

        # F - Разрядность (может быть дробной, например 2.5, 3.65)
        grade_value = row[5].value
        if grade_value is None:
            raise ValueError('Разрядность не указана')
        
        try:
            required_grade = Decimal(str(grade_value))
            if required_grade < 1 or required_grade > 5:
                raise ValueError('Разрядность должна быть от 1 до 5')
            
            # Находим ближайший целый разряд вниз для связи с WorkerGrade
            # (для базовой ставки, дробная часть сохраняется в required_grade)
            grade_int = int(required_grade)  # Округляем вниз, не вверх!
            if grade_int < 1:
                grade_int = 1
            elif grade_int >= 5:
                grade_int = 5
        except (InvalidOperation, ValueError) as e:
            raise ValueError(f'Некорректное значение разрядности: {grade_value}')

        # G - Комментарий (опционально)
        comment = ''
        if len(row) > 6 and row[6].value:
            comment = str(row[6].value).strip()

        # Находим базовый разряд (целый, для связи с WorkerGrade)
        try:
            grade = WorkerGrade.objects.get(grade=grade_int)
        except WorkerGrade.DoesNotExist:
            raise ValueError(f'Разряд {grade_int} не найден в справочнике')

        # Проверяем, существует ли работа с таким артикулом
        existing_work = WorkItem.objects.filter(article=article, is_current=True).first()

        if existing_work:
            # Создаём новую версию
            new_version = existing_work.create_new_version()
            new_version.section = section
            new_version.name = name
            new_version.unit = unit
            new_version.hours = hours
            new_version.grade = grade
            new_version.required_grade = required_grade
            new_version.composition = composition
            new_version.comment = comment
            new_version.coefficient = Decimal('1.00')  # По умолчанию
            return new_version
        else:
            # Создаём новую работу
            return WorkItem(
                article=article,
                section=section,
                name=name,
                unit=unit,
                hours=hours,
                grade=grade,
                required_grade=required_grade,
                composition=composition,
                comment=comment,
                coefficient=Decimal('1.00'),
                is_current=True,
                version_number=1
            )
