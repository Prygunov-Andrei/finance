"""
Excel invoice parser — извлекает данные из Excel-счетов через LLM (текстовый промпт).

Стратегия: openpyxl → текстовая таблица → LLM text prompt → ParsedInvoice.
Текстовый промпт ~10x дешевле, чем Vision для Excel-файлов.
"""
import hashlib
import json
import logging
import time
from typing import Tuple

import openpyxl

from ..models import LLMProvider, ParsedDocument
from ..providers import get_provider
from ..schemas import ParsedInvoice
from .exceptions import RateLimitError

logger = logging.getLogger(__name__)

# Лимиты для извлечения текста
MAX_ROWS = 200
MAX_COLS = 20


class ExcelInvoiceParser:
    """Парсит Excel-счета через LLM (текстовый промпт, без Vision)."""

    def __init__(self, provider_model: LLMProvider = None):
        self.provider_model = provider_model or LLMProvider.get_default()

    def parse(
        self, file_content: bytes, filename: str
    ) -> Tuple[ParsedInvoice, int]:
        """
        Парсит Excel-файл и возвращает структурированные данные.

        Args:
            file_content: содержимое .xlsx/.xls в байтах
            filename: имя файла (для логирования)

        Returns:
            (ParsedInvoice, processing_time_ms)
        """
        start = time.time()

        text = self._extract_text(file_content, filename)
        if not text.strip():
            raise ValueError(f'Не удалось извлечь текст из Excel-файла: {filename}')

        parsed_invoice = self._send_to_llm(text)

        elapsed_ms = int((time.time() - start) * 1000)
        return parsed_invoice, elapsed_ms

    def get_file_hash(self, file_content: bytes) -> str:
        return hashlib.sha256(file_content).hexdigest()

    # ------------------------------------------------------------------
    # Internal
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_text(file_content: bytes, filename: str) -> str:
        """Извлекает текст из Excel → форматирует как таблицу.

        Поддерживает .xlsx (openpyxl) и .xls (xlrd).
        """
        from pathlib import Path

        ext = Path(filename).suffix.lower()

        if ext == '.xls':
            return ExcelInvoiceParser._extract_text_xls(file_content, filename)
        return ExcelInvoiceParser._extract_text_xlsx(file_content, filename)

    @staticmethod
    def _extract_text_xlsx(file_content: bytes, filename: str) -> str:
        """Извлекает текст из .xlsx через openpyxl."""
        from io import BytesIO

        try:
            wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        except Exception as exc:
            raise ValueError(f'Не удалось открыть Excel-файл {filename}: {exc}')

        all_text_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_text = []

            for row_idx, row in enumerate(ws.iter_rows(max_row=MAX_ROWS, max_col=MAX_COLS, values_only=True), 1):
                cells = []
                for cell in row:
                    if cell is not None:
                        cells.append(str(cell).strip())
                    else:
                        cells.append('')

                if not any(cells):
                    continue

                rows_text.append(' | '.join(cells))

            if rows_text:
                if len(wb.sheetnames) > 1:
                    all_text_parts.append(f'=== Лист: {sheet_name} ===')
                all_text_parts.extend(rows_text)

        wb.close()
        return '\n'.join(all_text_parts)

    @staticmethod
    def _extract_text_xls(file_content: bytes, filename: str) -> str:
        """Извлекает текст из .xls (бинарный формат) через xlrd."""
        try:
            import xlrd
        except ImportError:
            raise ValueError(
                f'Для обработки .xls файлов необходима библиотека xlrd. '
                f'Установите: pip install xlrd'
            )

        try:
            wb = xlrd.open_workbook(file_contents=file_content)
        except Exception as exc:
            raise ValueError(f'Не удалось открыть .xls файл {filename}: {exc}')

        all_text_parts = []

        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            rows_text = []

            for row_idx in range(min(ws.nrows, MAX_ROWS)):
                cells = []
                for col_idx in range(min(ws.ncols, MAX_COLS)):
                    cell = ws.cell(row_idx, col_idx)
                    if cell.value is not None and str(cell.value).strip():
                        cells.append(str(cell.value).strip())
                    else:
                        cells.append('')

                if not any(cells):
                    continue

                rows_text.append(' | '.join(cells))

            if rows_text:
                if wb.nsheets > 1:
                    all_text_parts.append(f'=== Лист: {ws.name} ===')
                all_text_parts.extend(rows_text)

        return '\n'.join(all_text_parts)

    def _send_to_llm(self, text: str) -> ParsedInvoice:
        """Отправляет текст в LLM и получает ParsedInvoice."""
        system_prompt = """Ты — эксперт по распознаванию российских счетов на оплату.

Тебе предоставлен текст, извлечённый из Excel-файла со счётом.
Твоя задача — извлечь все данные и вернуть их в формате JSON.

Обязательные поля:
- vendor: информация о поставщике (name, inn, kpp)
- buyer: информация о покупателе (name, inn)
- invoice: номер и дата счёта (number, date в формате YYYY-MM-DD)
- totals: итоговые суммы (amount_gross — сумма с НДС, vat_amount — сумма НДС)
- items: массив позиций (name, quantity, unit, price_per_unit)
- confidence: твоя уверенность в корректности данных от 0.0 до 1.0

Правила:
1. Если не можешь определить значение — используй null
2. ИНН должен быть строкой из 10 или 12 цифр
3. Цены и суммы — десятичные числа
4. Единицы измерения: шт, м, м², м³, кг, т, л, компл, ч, усл, ед
5. Дата в формате YYYY-MM-DD

Верни ТОЛЬКО валидный JSON без markdown-форматирования."""

        user_prompt = f'Распарси этот счёт на оплату:\n\n{text}'

        raw_text = self._call_llm(system_prompt, user_prompt)
        return self._parse_response(raw_text)

    def _call_llm(self, system_prompt: str, user_prompt: str) -> str:
        """Вызывает LLM и возвращает сырой текст ответа."""
        provider_type = self.provider_model.provider_type
        api_key = self.provider_model.get_api_key()
        model_name = self.provider_model.model_name

        try:
            if provider_type == 'openai':
                import openai
                client = openai.OpenAI(api_key=api_key, timeout=120)
                response = client.chat.completions.create(
                    model=model_name,
                    messages=[
                        {'role': 'system', 'content': system_prompt},
                        {'role': 'user', 'content': user_prompt},
                    ],
                    temperature=0.1,
                    max_tokens=4000,
                )
                return response.choices[0].message.content.strip()

            elif provider_type == 'grok':
                import httpx
                with httpx.Client(timeout=120) as client:
                    resp = client.post(
                        'https://api.x.ai/v1/chat/completions',
                        headers={
                            'Authorization': f'Bearer {api_key}',
                            'Content-Type': 'application/json',
                        },
                        json={
                            'model': model_name,
                            'messages': [
                                {'role': 'system', 'content': system_prompt},
                                {'role': 'user', 'content': user_prompt},
                            ],
                            'temperature': 0.1,
                            'max_tokens': 4000,
                        },
                    )
                    resp.raise_for_status()
                    return resp.json()['choices'][0]['message']['content'].strip()

            elif provider_type == 'gemini':
                import httpx as _httpx
                url = (
                    f'https://generativelanguage.googleapis.com/v1beta'
                    f'/models/{model_name}:generateContent?key={api_key}'
                )
                payload = {
                    'contents': [{'parts': [{'text': f'{system_prompt}\n\n{user_prompt}'}]}],
                    'generationConfig': {'temperature': 0.1},
                }
                with _httpx.Client(timeout=120) as _client:
                    resp = _client.post(url, json=payload)
                    resp.raise_for_status()
                    return resp.json()['candidates'][0]['content']['parts'][0]['text'].strip()

            else:
                raise RuntimeError(f'Unsupported provider: {provider_type}')

        except Exception as exc:
            error_str = str(exc).lower()
            if '429' in error_str or 'rate limit' in error_str:
                raise RateLimitError(str(exc))
            raise

    @staticmethod
    def _parse_response(raw_text: str) -> ParsedInvoice:
        """Парсит JSON-ответ LLM → ParsedInvoice."""
        text = raw_text.strip()
        if text.startswith('```'):
            text = text.strip('`').strip()
            if text.startswith('json'):
                text = text[4:].strip()

        data = json.loads(text)
        return ParsedInvoice(**data)
