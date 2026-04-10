"""Генерация Excel-файла ТКП (Технического коммерческого предложения).

Использование:
    from proposals.services.tkp_excel_generator import TKPExcelGenerator
    buffer = TKPExcelGenerator(tkp).generate()
"""
import logging
from decimal import Decimal
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from proposals.models import TechnicalProposal

logger = logging.getLogger(__name__)


class TKPExcelGenerator:
    """Генерация multi-sheet ТКП в Excel."""

    def __init__(self, tkp: TechnicalProposal):
        self.tkp = tkp

    def generate(self) -> BytesIO:
        """Генерирует полный Excel-файл ТКП."""
        wb = openpyxl.Workbook()

        self._write_cover(wb.active)
        self._write_equipment(wb.create_sheet('Оборудование'))
        self._write_works(wb.create_sheet('Монтажные работы'))
        self._write_front_of_work(wb.create_sheet('Фронт работ'))
        self._write_terms(wb.create_sheet('Условия'))

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ── Стили ──────────────────────────────────────────────────

    @staticmethod
    def _styles():
        return {
            'title': Font(bold=True, size=16),
            'subtitle': Font(bold=True, size=12),
            'section': Font(bold=True, size=11, color='1F4E79'),
            'header': Font(bold=True, size=10),
            'bold': Font(bold=True),
            'center': Alignment(horizontal='center', vertical='center'),
            'wrap': Alignment(wrap_text=True, vertical='top'),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'),
            ),
            'num_fmt': '#,##0.00',
            'header_fill': PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
        }

    # ── Лист 1: Титульный ──────────────────────────────────────

    def _write_cover(self, ws):
        ws.title = 'ТКП'
        s = self._styles()

        ws.merge_cells('A1:F1')
        ws['A1'] = 'ТЕХНИЧЕСКОЕ КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ'
        ws['A1'].font = s['title']
        ws['A1'].alignment = s['center']

        ws.merge_cells('A3:F3')
        ws['A3'] = f'ТКП №{self.tkp.number}'
        ws['A3'].font = s['subtitle']
        ws['A3'].alignment = s['center']

        row = 5
        fields = [
            ('Дата', self.tkp.date.strftime('%d.%m.%Y') if self.tkp.date else ''),
            ('Объект', str(self.tkp.object)),
            ('Площадь объекта', f'{self.tkp.object_area} м²' if self.tkp.object_area else ''),
            ('Наша компания', str(self.tkp.legal_entity)),
            ('Название', self.tkp.name),
            ('Срок действия', f'{self.tkp.validity_days} дней'),
        ]
        for label, value in fields:
            ws.cell(row=row, column=1, value=label).font = s['bold']
            ws.cell(row=row, column=2, value=value)
            row += 1

        if self.tkp.advance_required:
            row += 1
            ws.cell(row=row, column=1, value='Необходимый аванс:').font = s['bold']
            row += 1
            ws.cell(row=row, column=1, value=self.tkp.advance_required)

        if self.tkp.work_duration:
            row += 1
            ws.cell(row=row, column=1, value='Сроки проведения работ:').font = s['bold']
            row += 1
            ws.cell(row=row, column=1, value=self.tkp.work_duration)

        if self.tkp.notes:
            row += 1
            ws.cell(row=row, column=1, value='Примечания:').font = s['bold']
            row += 1
            ws.cell(row=row, column=1, value=self.tkp.notes)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

    # ── Лист 2: Оборудование ──────────────────────────────────

    def _write_equipment(self, ws):
        s = self._styles()
        from estimates.models import EstimateItem

        ws.merge_cells('A1:H1')
        ws['A1'] = 'ОБОРУДОВАНИЕ И МАТЕРИАЛЫ'
        ws['A1'].font = s['title']
        ws['A1'].alignment = s['center']

        headers = ['№', 'Наименование', 'Модель', 'Ед.', 'Кол-во',
                   'Цена, ₽', 'Сумма, ₽']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = s['header']
            cell.alignment = s['center']
            cell.border = s['border']
            cell.fill = s['header_fill']

        row = 4
        grand_total = Decimal('0')

        for estimate in self.tkp.estimates.all():
            items = EstimateItem.objects.filter(
                estimate=estimate,
            ).select_related('section').order_by('section__sort_order', 'sort_order')

            current_section = None
            idx = 0

            for item in items:
                # Section header
                if item.section_id != current_section:
                    current_section = item.section_id
                    section_name = item.section.name if item.section else 'Без раздела'
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                    ws.cell(row=row, column=1, value=section_name).font = s['section']
                    row += 1

                idx += 1
                qty = item.quantity or Decimal('0')
                # Используем sale price если есть, иначе purchase price
                price = item.material_sale_unit_price or item.material_unit_price or Decimal('0')
                total = price * qty
                grand_total += total

                values = [idx, item.name, item.model_name or '', item.unit,
                          float(qty), float(price), float(total)]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = s['border']
                    if isinstance(val, float):
                        cell.number_format = s['num_fmt']
                row += 1

        # Grand total
        row += 1
        ws.cell(row=row, column=1, value='ИТОГО оборудование:').font = s['bold']
        ws.cell(row=row, column=7, value=float(grand_total)).number_format = s['num_fmt']
        ws.cell(row=row, column=7).font = s['bold']

        widths = [5, 35, 15, 6, 8, 12, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Лист 3: Монтажные работы ──────────────────────────────

    def _write_works(self, ws):
        s = self._styles()
        from estimates.models import EstimateItem

        ws.merge_cells('A1:G1')
        ws['A1'] = 'МОНТАЖНЫЕ РАБОТЫ'
        ws['A1'].font = s['title']
        ws['A1'].alignment = s['center']

        headers = ['№', 'Наименование оборудования', 'Работа', 'Часы',
                   'Кол-во', 'Стоимость работ, ₽']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = s['header']
            cell.alignment = s['center']
            cell.border = s['border']
            cell.fill = s['header_fill']

        row = 4
        grand_total = Decimal('0')
        total_hours = Decimal('0')

        for estimate in self.tkp.estimates.all():
            items = EstimateItem.objects.filter(
                estimate=estimate,
                work_item__isnull=False,
            ).select_related('work_item', 'section').order_by('section__sort_order', 'sort_order')

            idx = 0
            for item in items:
                idx += 1
                qty = item.quantity or Decimal('0')
                work_price = item.work_sale_unit_price or item.work_unit_price or Decimal('0')
                work_total = work_price * qty
                hours = Decimal(str(item.work_item.hours or 0)) * qty

                grand_total += work_total
                total_hours += hours

                values = [
                    idx, item.name,
                    item.work_item.name if item.work_item else '',
                    float(hours),
                    float(qty), float(work_total),
                ]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = s['border']
                    if isinstance(val, float):
                        cell.number_format = s['num_fmt']
                row += 1

        row += 1
        ws.cell(row=row, column=1, value='ИТОГО работы:').font = s['bold']
        ws.cell(row=row, column=4, value=float(total_hours)).font = s['bold']
        ws.cell(row=row, column=4).number_format = s['num_fmt']
        ws.cell(row=row, column=6, value=float(grand_total)).font = s['bold']
        ws.cell(row=row, column=6).number_format = s['num_fmt']

        widths = [5, 30, 25, 8, 8, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Лист 4: Фронт работ ──────────────────────────────────

    def _write_front_of_work(self, ws):
        s = self._styles()
        from proposals.models import TKPFrontOfWork

        ws.merge_cells('A1:B1')
        ws['A1'] = 'ФРОНТ РАБОТ'
        ws['A1'].font = s['title']
        ws['A1'].alignment = s['center']

        ws.cell(row=3, column=1, value='Для выполнения работ Заказчик обеспечивает:').font = s['subtitle']

        fow_items = TKPFrontOfWork.objects.filter(
            tkp=self.tkp,
        ).select_related('front_of_work_item').order_by('sort_order')

        row = 5
        for idx, item in enumerate(fow_items, 1):
            ws.cell(row=row, column=1, value=f'{idx}.').font = s['bold']
            ws.cell(row=row, column=2, value=item.front_of_work_item.name)
            ws.cell(row=row, column=2).alignment = s['wrap']
            row += 1

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 80

    # ── Лист 5: Условия ──────────────────────────────────────

    def _write_terms(self, ws):
        s = self._styles()

        ws.merge_cells('A1:B1')
        ws['A1'] = 'УСЛОВИЯ'
        ws['A1'].font = s['title']
        ws['A1'].alignment = s['center']

        row = 3
        terms = [
            ('Срок действия предложения', f'{self.tkp.validity_days} дней с даты ТКП'),
            ('Необходимый аванс', self.tkp.advance_required or 'По согласованию'),
            ('Сроки выполнения работ', self.tkp.work_duration or 'По согласованию'),
        ]

        for label, value in terms:
            ws.cell(row=row, column=1, value=label).font = s['bold']
            ws.cell(row=row, column=2, value=value).alignment = s['wrap']
            row += 1

        row += 2
        ws.cell(row=row, column=1, value='Подписи:').font = s['subtitle']
        row += 2
        ws.cell(row=row, column=1, value='От Исполнителя:')
        ws.cell(row=row, column=2, value='________________________')
        row += 2
        ws.cell(row=row, column=1, value='От Заказчика:')
        ws.cell(row=row, column=2, value='________________________')

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
